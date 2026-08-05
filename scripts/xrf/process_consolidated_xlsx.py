"""
Convert EQ_XRF_All_Trips_Consolidated.xlsx (Trips 1-4) into the TSV schema
already used by data/processed/geochemistry/xrf_lab_table_filtered.tsv.

Each sheet (T1-Dr, T1-Sr, ..., T4-PRr) contributes rows. For every
(Trip, Site, Compartment) we emit ONE row with all element + oxide
concentrations populated from across the multiple XRF instrument runs
(XRF 0 = Elements mode, XRF 1 = Oxides mode, sometimes XRF 2/3 = re-runs).
For an analyte that appears in more than one run, we keep the largest
non-zero value (XRF non-detects come back as 0 / below LLD; the highest
reading is the most informative).

Sample ID convention (matches existing samples ABox):
    Trip 1 → no prefix          e.g. 10Dr1
    Trip 2 → "T" prefix         e.g. T2Dr1
    Trip 3 → "F" prefix         e.g. F10Dr1
    Trip 4 → "S" prefix         e.g. S10Dr1
We always tag rep "1" because the existing Trip-5 lab data uses rep 1 too —
the XRF analytical replicates per site map to the same physical sample.

Output schema preserves the column order of the existing
xrf_lab_table_filtered.tsv exactly so generate_xrf_abox.groovy needs
no changes.

Usage:
    uv run python scripts/xrf/process_consolidated_xlsx.py \\
        data/metadata/xrf/all-trips-consolidated/EQ_XRF_All_Trips_Consolidated.xlsx \\
        data/processed/geochemistry/xrf_lab_table_trips1-4.tsv
"""

from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl


TRIP_PREFIX = {1: "", 2: "T", 3: "F", 4: "S"}
COMPARTMENT_TO_CODE = {"Deep": "Dr", "Surface": "Sr", "Rhizosphere": "PRr"}
COMPARTMENT_TO_SOIL = {"Deep": "Deep", "Surface": "Surface", "Rhizosphere": "Rhizosphere"}

# Header order from data/processed/geochemistry/xrf_lab_table_filtered.tsv —
# kept verbatim so generate_xrf_abox.groovy can read both files identically.
HEADER = [
    "SampleID", "SoilType",
    "Al", "Al2O3", "Ba", "BaO", "Br",
    "Ca", "CaO", "Ce", "CeO2", "Cl",
    "Co", "CoO", "Cr", "Cr2O3", "Cs",
    "Cu", "CuO",
    "Diameter",
    "Eu", "Eu2O3",
    "Fe", "Fe2O3",
    "Ga", "Ga2O3",
    "Gd", "Gd2O3",
    "K", "K2O",
    "La", "La2O3",
    "Material", "Method",
    "Mg", "MgO",
    "Mn", "MnO",
    "Mo", "MoO3",
    "Mode",
    "Na", "Na2O",
    "Nd", "Nd2O3",
    "Ni", "NiO",
    "P", "P2O5",
    "Pr", "Pr6O11",
    "S", "SO3",
    "Sc", "Sc2O3",
    "Si", "SiO2",
    "Sm", "Sm2O3",
    "Sr", "SrO",
    "Ti", "TiO2",
    "V", "V2O5",
    "Zn", "ZnO",
    "Zr", "ZrO2",
]
META_COLS = {"SampleID", "SoilType", "Diameter", "Material", "Method", "Mode"}
ANALYTE_COLS = [c for c in HEADER if c not in META_COLS]


def parse_sheet_name(name: str) -> tuple[int, str] | None:
    # Sheet pattern: T<trip>-<compartmentCode>, e.g. T1-Dr, T2-PRr.
    if not name.startswith("T"):
        return None
    try:
        trip = int(name[1])
    except ValueError:
        return None
    code = name.split("-", 1)[1] if "-" in name else ""
    code_to_compartment = {v: k for k, v in COMPARTMENT_TO_CODE.items()}
    compartment = code_to_compartment.get(code)
    if compartment is None:
        return None
    return trip, compartment


def collect_sheet(ws, trip: int, compartment: str) -> dict[int, dict[str, float]]:
    """Return {site_no: {analyte: max_non_zero_concentration}}."""
    data: dict[int, dict[str, float]] = defaultdict(dict)
    trip_label = f"Trip {trip}"
    for row in ws.iter_rows(values_only=True):
        if len(row) < 7:
            continue
        if row[0] != trip_label:
            continue  # skip headers, dividers, banner rows
        site = row[2]
        formula = row[3]
        conc = row[5]
        if site is None or formula is None or conc is None:
            continue
        try:
            conc_f = float(conc)
        except (TypeError, ValueError):
            continue
        if conc_f <= 0:
            continue  # below LLD / not detected
        if formula not in ANALYTE_COLS:
            # E.g. As, Pr2O3 — analytes the existing TSV doesn't carry.
            # We silently drop them here; track at scan time if needed.
            continue
        prev = data[site].get(formula, 0.0)
        if conc_f > prev:
            data[site][formula] = conc_f
    return data


def emit_rows(
    xlsx_path: Path,
) -> tuple[list[list[str]], dict[str, int]]:
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    rows: list[list[str]] = []
    skipped_analytes: dict[str, int] = defaultdict(int)
    sheets_processed = 0
    for sheet_name in wb.sheetnames:
        meta = parse_sheet_name(sheet_name)
        if meta is None:
            continue
        trip, compartment = meta
        ws = wb[sheet_name]
        # Quick second pass to record analytes we skipped (not in HEADER).
        trip_label = f"Trip {trip}"
        for row in ws.iter_rows(values_only=True):
            if len(row) < 7 or row[0] != trip_label:
                continue
            formula = row[3]
            if formula and formula not in ANALYTE_COLS:
                skipped_analytes[formula] += 1
        site_data = collect_sheet(ws, trip, compartment)
        prefix = TRIP_PREFIX[trip]
        compartment_code = COMPARTMENT_TO_CODE[compartment]
        soil_type = COMPARTMENT_TO_SOIL[compartment]
        for site_no in sorted(site_data.keys()):
            sample_id = f"{prefix}{site_no}{compartment_code}1"
            row_values = []
            for col in HEADER:
                if col == "SampleID":
                    row_values.append(sample_id)
                elif col == "SoilType":
                    row_values.append(soil_type)
                elif col in {"Diameter", "Material", "Method", "Mode"}:
                    row_values.append("")  # not provided by this xlsx
                else:
                    v = site_data[site_no].get(col, 0.0)
                    row_values.append(f"{v:g}")
            rows.append(row_values)
        sheets_processed += 1
    print(f"  processed {sheets_processed} sheets, emitted {len(rows)} rows", file=sys.stderr)
    return rows, dict(skipped_analytes)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("out_tsv", type=Path)
    args = parser.parse_args()

    if not args.xlsx.exists():
        print(f"ERROR: xlsx not found: {args.xlsx}", file=sys.stderr)
        return 1

    rows, skipped = emit_rows(args.xlsx)
    args.out_tsv.parent.mkdir(parents=True, exist_ok=True)
    with args.out_tsv.open("w") as f:
        f.write("\t".join(HEADER) + "\n")
        for r in rows:
            f.write("\t".join(r) + "\n")

    if skipped:
        print(
            "  analytes present in xlsx but absent from output schema "
            "(silently dropped):",
            file=sys.stderr,
        )
        for k, n in sorted(skipped.items(), key=lambda x: -x[1]):
            print(f"    {k}: {n} measurements", file=sys.stderr)

    print(f"  wrote {args.out_tsv} ({len(rows)} samples)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
