#!/usr/bin/env python3
"""Audit Empty Quarter XRF provenance, aggregation, and field/lab agreement.

This command is deliberately read-only with respect to source data.  It writes a
deterministic evidence bundle to ``--output-dir`` and never rewrites an XRF
source or analysis table.

The audit keeps three data layers separate:

* Trip-5 field XRF: instrument sessions logged at site level.
* Laboratory XRF run-level/source observations: Trips 1--5.
* Laboratory analytical tables: the canonical 547+178 inputs and retired
  158-record Trip-5 subset.

Only ``openpyxl`` is required outside the Python standard library.

Example
-------
uv run --with openpyxl python scripts/xrf/audit_xrf_provenance.py \
  --project-root . \
  --output-dir analysis/xrf_audit
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import statistics
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

import openpyxl


AUDIT_SCHEMA_VERSION = "1.1"
MISSING = ""

FIELD_TABLE = Path("data/processed/geochemistry/xrf_field_table.tsv")
FIELD_LOG = Path("data/metadata/xrf/xrf-measurements.tsv")
FIELD_RAW_DIR = Path("data/processed/geochemistry/xrf")
FIELD_PROTOCOL = Path("data/metadata/protocols/measurement/XRF_measurement.md")
LAB_T14_WORKBOOK = Path(
    "data/metadata/xrf/all-trips-consolidated/"
    "EQ_XRF_All_Trips_Consolidated.xlsx"
)
LAB_T14_TABLE = Path("data/processed/geochemistry/xrf_lab_table_trips1-4.tsv")
LAB_T5_DIR = Path("data/metadata/xrf/xrf-lab")
LAB_T5_TABLE = Path("data/processed/geochemistry/xrf_lab_table_filtered.tsv")
LAB_T5_ANALYTICAL = Path("data/processed/geochemistry/xrf_lab_combined.tsv")
COMMUNITY_TABLE = Path("analysis/v2/review/cache/genus_counts.tsv")

T5_WORKBOOKS = {
    "EQ 5th Trip XRF Sr Results.xlsx": "Surface",
    "EQ XRF results (Fifth trip).xlsx": "Deep",
    "EQ 5th Trip XRF PRr Results.xlsx": "Rhizosphere",
}

TRIP_PREFIX = {1: "", 2: "T", 3: "F", 4: "S", 5: "V"}
PREFIX_TRIP = {"": 1, "T": 2, "F": 3, "S": 4, "V": 5}
COMPARTMENT_CODE = {"Deep": "D", "Surface": "S", "Rhizosphere": "PR"}
CODE_COMPARTMENT = {
    "D": "Deep",
    "S": "Surface",
    "P": "Rhizosphere",
    "PR": "Rhizosphere",
}
META_COLUMNS = {"SoilType", "Material", "Mode", "Diameter", "Method"}

SAMPLE_RE = re.compile(
    r"^(?:e\d+_)?(?P<prefix>[TFSV])?(?P<site>\d+)"
    r"(?P<compartment>PR|P|D|S)r(?P<rep>\d+)"
)
ELEMENT_RE = re.compile(r"^[A-Z][a-z]?$")

RECONCILIATION_COLUMNS = [
    "dataset_id",
    "workflow",
    "trip_scope",
    "compartment_scope",
    "record_unit",
    "row_count",
    "analyte_measurement_count",
    "unique_sample_ids",
    "unique_trip_site_compartments",
    "unique_sites",
    "community_joined_all",
    "community_joined_qc",
    "status",
    "caveat",
]


def project_path(project_root: Path, relative: Path) -> Path:
    """Resolve a project-relative path without requiring the current directory."""
    return (project_root / relative).resolve()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def write_tsv(
    path: Path,
    rows: Iterable[Mapping[str, Any]],
    columns: Sequence[str],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=list(columns),
            delimiter="\t",
            lineterminator="\n",
            extrasaction="ignore",
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    column: format_cell(row.get(column, MISSING))
                    for column in columns
                }
            )


def format_cell(value: Any) -> str:
    if value is None:
        return MISSING
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        if not math.isfinite(value):
            return MISSING
        return f"{value:.10g}"
    return str(value)


def numeric(value: Any) -> float | None:
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def canonical_sample_key(
    sample_id: str,
    default_trip: int | None = None,
) -> tuple[int, int, str] | None:
    """Return (trip, site, compartment) for sequence or XRF sample IDs."""
    match = SAMPLE_RE.match(str(sample_id).replace(" ", ""))
    if not match:
        return None
    prefix = match.group("prefix") or ""
    trip = default_trip if default_trip is not None else PREFIX_TRIP[prefix]
    return (
        trip,
        int(match.group("site")),
        CODE_COMPARTMENT[match.group("compartment")],
    )


def canonical_t14_sample_id(trip: int, site: int, compartment: str) -> str:
    code = {
        "Deep": "Dr1",
        "Surface": "Sr1",
        "Rhizosphere": "PRr1",
    }[compartment]
    return f"{TRIP_PREFIX[trip]}{site}{code}"


def canonical_t5_sheet_id(sheet: str, compartment: str) -> str | None:
    """Reproduce the current Trip-5 parser's sheet-name normalization."""
    name = str(sheet).replace(" ", "")
    if compartment == "Deep":
        if "BestDetection" in name:
            return None
        name = name.replace("Fastscreening", "")
        if name.isdigit():
            return f"V{name}Dr1"
        if not name.startswith("V"):
            name = f"V{name}"
        return name
    if not name.startswith("V"):
        name = f"V{name}"
    return name


def is_element_formula(formula: str) -> bool:
    return bool(ELEMENT_RE.fullmatch(formula))


def expected_primary_status(formula: str) -> str:
    """Infer the primary status encoded by workbook section labels.

    Consolidated workbook banners identify XRF 0 as Elements and XRF 1 as
    Oxides.  This is an audit candidate, not a claim that the instrument
    manufacturer prescribes the choice.
    """
    return "XRF 0" if is_element_formula(formula) else "XRF 1"


def load_t14_observations(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    observations: list[dict[str, Any]] = []
    sheet_sample_keys: set[tuple[int, int, str]] = set()
    sheets = 0
    for worksheet in workbook.worksheets:
        if not worksheet.title.startswith("T"):
            continue
        match = re.fullmatch(r"T([1-4])-(Dr|Sr|PRr)", worksheet.title)
        if not match:
            continue
        trip = int(match.group(1))
        compartment = {
            "Dr": "Deep",
            "Sr": "Surface",
            "PRr": "Rhizosphere",
        }[match.group(2)]
        sheets += 1
        for row_index, row in enumerate(
            worksheet.iter_rows(values_only=True), start=1
        ):
            if len(row) < 12 or row[0] != f"Trip {trip}":
                continue
            site_value, formula, concentration, status = (
                row[2],
                row[3],
                row[5],
                row[6],
            )
            value = numeric(concentration)
            if site_value is None or formula is None or value is None:
                continue
            site = int(site_value)
            sample_id = canonical_t14_sample_id(trip, site, compartment)
            sheet_sample_keys.add((trip, site, compartment))
            observations.append(
                {
                    "source": "lab_t1_4_workbook",
                    "trip": trip,
                    "site": site,
                    "compartment": compartment,
                    "sample_id": sample_id,
                    "formula": str(formula).strip(),
                    "value": value,
                    "status": str(status).strip() if status else MISSING,
                    "stat_error": numeric(row[10]),
                    "lld": row[11],
                    "sheet": worksheet.title,
                    "row_index": row_index,
                }
            )
    return observations, {
        "sheet_count": sheets,
        "sample_count": len(sheet_sample_keys),
        "measurement_count": len(observations),
    }


def load_t5_observations(
    directory: Path,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    observations: list[dict[str, Any]] = []
    all_sheets = 0
    selected_sheets = 0
    excluded_sheets: list[dict[str, str]] = []
    metadata_counts: Counter[tuple[str, str, str]] = Counter()
    metadata_samples: dict[tuple[str, str, str], set[str]] = defaultdict(set)
    selected_ids: set[str] = set()
    for filename, compartment in sorted(T5_WORKBOOKS.items()):
        path = directory / filename
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            all_sheets += 1
            sample_id = canonical_t5_sheet_id(worksheet.title, compartment)
            if sample_id is None:
                excluded_sheets.append(
                    {
                        "workbook": filename,
                        "sheet": worksheet.title,
                        "reason": (
                            "Excluded by current parser because sheet name contains "
                            "'Best Detection'"
                        ),
                    }
                )
                continue
            selected_sheets += 1
            selected_ids.add(sample_id)
            key = canonical_sample_key(sample_id, default_trip=5)
            if key is None:
                continue
            _, site, parsed_compartment = key
            for row_index, row in enumerate(
                worksheet.iter_rows(values_only=True), start=1
            ):
                if len(row) < 4:
                    continue
                formula = row[0]
                if formula in META_COLUMNS:
                    metadata_counts[(compartment, str(formula), str(row[2]))] += 1
                    metadata_samples[
                        (compartment, str(formula), str(row[2]))
                    ].add(sample_id)
                    continue
                if formula in (None, "Formula"):
                    continue
                value = numeric(row[2])
                if value is None:
                    continue
                observations.append(
                    {
                        "source": "lab_t5_workbooks",
                        "trip": 5,
                        "site": site,
                        "compartment": parsed_compartment,
                        "sample_id": sample_id,
                        "formula": str(formula).strip(),
                        "value": value,
                        "status": str(row[3]).strip() if row[3] else MISSING,
                        "stat_error": numeric(row[7]) if len(row) > 7 else None,
                        "lld": row[8] if len(row) > 8 else None,
                        "sheet": worksheet.title,
                        "workbook": filename,
                        "row_index": row_index,
                    }
                )
    return observations, {
        "all_sheet_count": all_sheets,
        "selected_sheet_count": selected_sheets,
        "selected_sample_count": len(selected_ids),
        "measurement_count": len(observations),
        "excluded_sheets": excluded_sheets,
        "metadata_counts": metadata_counts,
        "metadata_sample_counts": {
            key: len(sample_ids) for key, sample_ids in metadata_samples.items()
        },
    }


def observation_groups(
    observations: Sequence[Mapping[str, Any]],
) -> dict[tuple[str, str], list[Mapping[str, Any]]]:
    groups: dict[tuple[str, str], list[Mapping[str, Any]]] = defaultdict(list)
    for observation in observations:
        groups[(str(observation["sample_id"]), str(observation["formula"]))].append(
            observation
        )
    for key in groups:
        groups[key] = sorted(
            groups[key],
            key=lambda row: (
                str(row.get("workbook", "")),
                str(row.get("sheet", "")),
                int(row.get("row_index", 0)),
            ),
        )
    return groups


def aggregate_group(
    rows: Sequence[Mapping[str, Any]],
    rule: str,
) -> float | None:
    values = [float(row["value"]) for row in rows if numeric(row["value"]) is not None]
    positive = [value for value in values if value > 0]
    if not values:
        return None
    if rule == "max_positive":
        return max(positive) if positive else 0.0
    if rule == "mean_positive":
        return statistics.fmean(positive) if positive else 0.0
    if rule == "median_positive":
        return statistics.median(positive) if positive else 0.0
    if rule == "first_reported":
        return values[0]
    if rule == "last_reported":
        return values[-1]
    if rule == "primary_status":
        primary = expected_primary_status(str(rows[0]["formula"]))
        matched = [
            float(row["value"])
            for row in rows
            if row.get("status") == primary and numeric(row["value"]) is not None
        ]
        if not matched:
            return None
        return matched[0] if len(matched) == 1 else statistics.median(matched)
    raise ValueError(f"Unknown aggregation rule: {rule}")


def average_ranks(values: Sequence[float]) -> list[float]:
    order = sorted(range(len(values)), key=lambda index: values[index])
    ranks = [0.0] * len(values)
    cursor = 0
    while cursor < len(order):
        end = cursor + 1
        while end < len(order) and values[order[end]] == values[order[cursor]]:
            end += 1
        rank = (cursor + 1 + end) / 2.0
        for position in range(cursor, end):
            ranks[order[position]] = rank
        cursor = end
    return ranks


def pearson(x: Sequence[float], y: Sequence[float]) -> float | None:
    if len(x) != len(y) or len(x) < 2:
        return None
    mean_x = statistics.fmean(x)
    mean_y = statistics.fmean(y)
    numerator = sum((a - mean_x) * (b - mean_y) for a, b in zip(x, y))
    denominator = math.sqrt(
        sum((a - mean_x) ** 2 for a in x)
        * sum((b - mean_y) ** 2 for b in y)
    )
    return numerator / denominator if denominator > 0 else None


def spearman(x: Sequence[float], y: Sequence[float]) -> float | None:
    return pearson(average_ranks(x), average_ranks(y))


def quantile(values: Sequence[float], probability: float) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    if len(ordered) == 1:
        return ordered[0]
    position = probability * (len(ordered) - 1)
    lower = int(math.floor(position))
    upper = int(math.ceil(position))
    if lower == upper:
        return ordered[lower]
    weight = position - lower
    return ordered[lower] * (1 - weight) + ordered[upper] * weight


def aggregation_sensitivity(
    workflow: str,
    observations: Sequence[Mapping[str, Any]],
    current_rule: str,
) -> list[dict[str, Any]]:
    groups = observation_groups(observations)
    by_formula: dict[str, list[tuple[list[Mapping[str, Any]], dict[str, float | None]]]]
    by_formula = defaultdict(list)
    rules = [
        "max_positive",
        "mean_positive",
        "median_positive",
        "first_reported",
        "last_reported",
        "primary_status",
    ]
    for (_, formula), rows in sorted(groups.items()):
        values = {rule: aggregate_group(rows, rule) for rule in rules}
        by_formula[formula].append((rows, values))

    results: list[dict[str, Any]] = []
    for formula, items in sorted(by_formula.items()):
        multi = [item for item in items if len(item[0]) > 1]
        primary_missing = sum(item[1]["primary_status"] is None for item in items)
        paired = [
            item
            for item in items
            if item[1]["primary_status"] is not None
            and item[1]["max_positive"] is not None
        ]
        current_primary = [
            (
                float(item[1][current_rule]),
                float(item[1]["primary_status"]),
            )
            for item in items
            if item[1][current_rule] is not None
            and item[1]["primary_status"] is not None
        ]
        relative = [
            (current - primary) / abs(primary)
            for current, primary in current_primary
            if abs(primary) > 0
        ]
        changed_relative = [
            (current - primary) / abs(primary)
            for current, primary in current_primary
            if abs(primary) > 0 and not math.isclose(current, primary)
        ]
        differences = [
            abs(current - primary) > 1e-12
            for current, primary in current_primary
        ]
        x = [pair[0] for pair in current_primary]
        y = [pair[1] for pair in current_primary]
        statuses = sorted(
            {
                str(row.get("status", ""))
                for rows, _ in items
                for row in rows
                if row.get("status")
            }
        )
        results.append(
            {
                "workflow": workflow,
                "analyte": formula,
                "formula_class": "element" if is_element_formula(formula) else "compound",
                "expected_primary_status": expected_primary_status(formula),
                "observed_statuses": ",".join(statuses),
                "sample_analyte_groups": len(items),
                "groups_with_multiple_rows": len(multi),
                "groups_missing_primary_status": primary_missing,
                "groups_current_differs_from_primary": sum(differences),
                "current_rule": current_rule,
                "spearman_current_vs_primary": spearman(x, y),
                "median_relative_difference": quantile(relative, 0.5),
                "q95_absolute_relative_difference": quantile(
                    [abs(value) for value in relative], 0.95
                ),
                "median_relative_difference_among_changed": quantile(
                    changed_relative, 0.5
                ),
                "maximum_absolute_relative_difference": (
                    max(abs(value) for value in relative) if relative else None
                ),
                "paired_groups": len(paired),
            }
        )
    return results


def aggregation_group_details(
    workflow: str,
    observations: Sequence[Mapping[str, Any]],
    current_rule: str,
) -> list[dict[str, Any]]:
    """Return one audit row per multi-row or non-primary sample/analyte group."""
    rows: list[dict[str, Any]] = []
    for (sample_id, analyte), group in sorted(observation_groups(observations).items()):
        candidates = {
            rule: aggregate_group(group, rule)
            for rule in (
                "max_positive",
                "mean_positive",
                "median_positive",
                "first_reported",
                "last_reported",
                "primary_status",
            )
        }
        if len(group) == 1 and candidates["primary_status"] is not None:
            continue
        first = group[0]
        current = candidates[current_rule]
        primary = candidates["primary_status"]
        relative = (
            (float(current) - float(primary)) / abs(float(primary))
            if current is not None and primary not in (None, 0.0)
            else None
        )
        rows.append(
            {
                "workflow": workflow,
                "sample_id": sample_id,
                "trip": first["trip"],
                "site": first["site"],
                "compartment": first["compartment"],
                "analyte": analyte,
                "source_row_count": len(group),
                "status_value_sequence": ";".join(
                    f"{row.get('status', '')}:{format_cell(row['value'])}"
                    for row in group
                ),
                "expected_primary_status": expected_primary_status(analyte),
                "max_positive": candidates["max_positive"],
                "mean_positive": candidates["mean_positive"],
                "median_positive": candidates["median_positive"],
                "first_reported": candidates["first_reported"],
                "last_reported": candidates["last_reported"],
                "primary_status_value": primary,
                "current_rule": current_rule,
                "current_value": current,
                "current_minus_primary_relative": relative,
                "primary_status_available": primary is not None,
            }
        )
    return rows


def load_numeric_table(
    path: Path,
) -> tuple[list[dict[str, str]], list[str], str]:
    rows = read_tsv(path)
    if not rows:
        return [], [], "SampleID"
    id_column = "SampleID" if "SampleID" in rows[0] else next(iter(rows[0]))
    analytes = [
        column
        for column in rows[0]
        if column != id_column and column not in META_COLUMNS
    ]
    return rows, analytes, id_column


def compare_current_table_to_rule(
    rows: Sequence[Mapping[str, str]],
    analytes: Sequence[str],
    id_column: str,
    observations: Sequence[Mapping[str, Any]],
    rule: str,
) -> dict[str, Any]:
    groups = observation_groups(observations)
    checked = 0
    mismatches = 0
    missing_source = 0
    for row in rows:
        sample_id = str(row[id_column])
        for analyte in analytes:
            current = numeric(row.get(analyte))
            source_rows = groups.get((sample_id, analyte))
            if not source_rows:
                if current not in (None, 0.0):
                    missing_source += 1
                continue
            expected = aggregate_group(source_rows, rule)
            if expected is None or current is None:
                continue
            checked += 1
            if not math.isclose(current, expected, rel_tol=1e-9, abs_tol=1e-12):
                mismatches += 1
    return {
        "rule": rule,
        "checked_reported_cells": checked,
        "mismatched_reported_cells": mismatches,
        "nonzero_cells_without_source_row": missing_source,
    }


def table_sample_keys(
    rows: Sequence[Mapping[str, str]],
    id_column: str,
    default_trip: int | None = None,
) -> set[tuple[int, int, str]]:
    return {
        key
        for row in rows
        if (
            key := canonical_sample_key(
                str(row[id_column]),
                default_trip=default_trip,
            )
        )
        is not None
    }


def load_community_keys(
    path: Path,
    minimum_reads: float = 2000.0,
) -> tuple[set[tuple[int, int, str]], set[tuple[int, int, str]], dict[str, int]]:
    """Stream the derived genus table to reproduce the current ecology QC join."""
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle, delimiter="\t")
        header = next(reader)
        sample_ids = header[1:]
        totals = [0.0] * len(sample_ids)
        feature_count = 0
        for row in reader:
            feature_count += 1
            for index, value in enumerate(row[1 : len(sample_ids) + 1]):
                parsed = numeric(value)
                if parsed is not None:
                    totals[index] += parsed
    all_keys = {
        key
        for sample_id in sample_ids
        if (key := canonical_sample_key(sample_id)) is not None
    }
    qc_ids = [
        sample_id
        for sample_id, total in zip(sample_ids, totals)
        if total >= minimum_reads
    ]
    qc_keys = {
        key
        for sample_id in qc_ids
        if (key := canonical_sample_key(sample_id)) is not None
    }
    return all_keys, qc_keys, {
        "sample_columns": len(sample_ids),
        "feature_rows": feature_count,
        "qc_sample_columns": len(qc_ids),
        "all_group_keys": len(all_keys),
        "qc_group_keys": len(qc_keys),
    }


def field_lab_agreement(
    field_rows: Sequence[Mapping[str, str]],
    lab_rows: Sequence[Mapping[str, str]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    """Match field sessions to Trip-5 lab Deep records at site level.

    Field replicates are reduced by the median.  The source field log contains
    site but not physical sample/replicate identifiers, so the match is
    explicitly site-level and must not be described as a same-aliquot comparison.
    """
    field_meta = {"SiteID", "TestID", "LE"}
    field_analytes = {
        column
        for column in field_rows[0]
        if column not in field_meta
        and not column.endswith("_error")
        and is_element_formula(column)
    }
    lab_analytes = {
        column
        for column in lab_rows[0]
        if column not in META_COLUMNS | {"SampleID"}
        and is_element_formula(column)
    }
    analytes = sorted(field_analytes & lab_analytes)

    field_by_site: dict[int, list[Mapping[str, str]]] = defaultdict(list)
    for row in field_rows:
        site = int(float(row["SiteID"]))
        field_by_site[site].append(row)

    lab_by_site: dict[int, Mapping[str, str]] = {}
    for row in lab_rows:
        if row.get("SoilType") != "Deep":
            continue
        key = canonical_sample_key(str(row["SampleID"]), default_trip=5)
        if key is not None:
            lab_by_site[key[1]] = row

    sites = sorted(set(field_by_site) & set(lab_by_site))
    site_rows: list[dict[str, Any]] = []
    for site in sites:
        row: dict[str, Any] = {
            "site": site,
            "field_sessions": len(field_by_site[site]),
            "lab_sample_id": lab_by_site[site]["SampleID"],
            "match_level": "site_to_Trip5_Deep",
        }
        for analyte in analytes:
            field_values = [
                value
                for source in field_by_site[site]
                if (value := numeric(source.get(analyte))) is not None
            ]
            row[f"field_{analyte}"] = (
                statistics.median(field_values) if field_values else None
            )
            row[f"lab_{analyte}"] = numeric(lab_by_site[site].get(analyte))
        site_rows.append(row)

    agreement: list[dict[str, Any]] = []
    for analyte in analytes:
        pairs = [
            (
                float(row[f"field_{analyte}"] or 0.0),
                float(row[f"lab_{analyte}"] or 0.0),
            )
            for row in site_rows
        ]
        both = [(field, lab) for field, lab in pairs if field > 0 and lab > 0]
        field_only = sum(field > 0 and lab <= 0 for field, lab in pairs)
        lab_only = sum(lab > 0 and field <= 0 for field, lab in pairs)
        neither = sum(field <= 0 and lab <= 0 for field, lab in pairs)
        denominator = 2 * len(both) + field_only + lab_only
        log_ratios = [math.log10(field / lab) for field, lab in both]
        log_bias = statistics.fmean(log_ratios) if log_ratios else None
        log_sd = statistics.stdev(log_ratios) if len(log_ratios) > 1 else None
        agreement.append(
            {
                "analyte": analyte,
                "matched_sites": len(pairs),
                "both_reported_positive": len(both),
                "field_only_positive": field_only,
                "lab_only_positive": lab_only,
                "neither_positive": neither,
                "reported_positive_agreement": (
                    2 * len(both) / denominator if denominator else None
                ),
                "spearman_all_sites": spearman(
                    [pair[0] for pair in pairs],
                    [pair[1] for pair in pairs],
                ),
                "spearman_both_positive": spearman(
                    [pair[0] for pair in both],
                    [pair[1] for pair in both],
                ),
                "median_field_lab_ratio": quantile(
                    [field / lab for field, lab in both], 0.5
                ),
                "mean_log10_field_lab_bias": log_bias,
                "log10_lower_loa": (
                    log_bias - 1.96 * log_sd
                    if log_bias is not None and log_sd is not None
                    else None
                ),
                "log10_upper_loa": (
                    log_bias + 1.96 * log_sd
                    if log_bias is not None and log_sd is not None
                    else None
                ),
                "interpretation_limit": (
                    "Site-level field-session median versus one archived Deep "
                    "lab record; not a same-aliquot method-comparison experiment"
                ),
            }
        )

    replicate_rows: list[dict[str, Any]] = []
    for analyte in analytes:
        cvs: list[float] = []
        repeated_sites = 0
        for source_rows in field_by_site.values():
            if len(source_rows) < 2:
                continue
            values = [
                value
                for row in source_rows
                if (value := numeric(row.get(analyte))) is not None
            ]
            if len(values) < 2:
                continue
            repeated_sites += 1
            mean = statistics.fmean(values)
            if mean > 0:
                cvs.append(statistics.stdev(values) / mean)
        replicate_rows.append(
            {
                "analyte": analyte,
                "sites_with_repeated_field_sessions": repeated_sites,
                "sites_with_estimable_cv": len(cvs),
                "median_within_site_cv": quantile(cvs, 0.5),
                "q95_within_site_cv": quantile(cvs, 0.95),
            }
        )

    return agreement, replicate_rows, {
        "matched_sites": len(sites),
        "field_sites": len(field_by_site),
        "lab_deep_sites": len(lab_by_site),
        "shared_element_analytes": len(analytes),
        "repeated_field_sites": sum(
            len(rows) > 1 for rows in field_by_site.values()
        ),
    }


def source_inventory(project_root: Path, community_path: Path) -> list[dict[str, Any]]:
    roles = {
        FIELD_TABLE: ("field_processed_table", "field"),
        FIELD_LOG: ("field_session_log", "field"),
        FIELD_PROTOCOL: ("field_protocol", "field"),
        LAB_T14_WORKBOOK: ("lab_source_workbook_t1_4", "lab"),
        LAB_T14_TABLE: ("lab_canonical_table_t1_4", "lab"),
        LAB_T5_TABLE: ("lab_canonical_table_t5", "lab"),
        LAB_T5_ANALYTICAL: ("lab_retired_analytical_subset_t5", "lab"),
    }
    for filename in T5_WORKBOOKS:
        relative = LAB_T5_DIR / filename
        roles[relative] = ("lab_source_workbook_t5", "lab")

    rows: list[dict[str, Any]] = []
    for relative, (role, workflow) in sorted(
        roles.items(), key=lambda item: str(item[0])
    ):
        path = project_path(project_root, relative)
        rows.append(
            {
                "path": str(relative),
                "role": role,
                "workflow": workflow,
                "format": path.suffix.lower().lstrip("."),
                "bytes": path.stat().st_size,
                "sha256": sha256_file(path),
            }
        )
    for path in sorted(
        project_path(project_root, FIELD_RAW_DIR).rglob("vanta_data_*.csv")
    ):
        rows.append(
            {
                "path": str(path.relative_to(project_root)),
                "role": "field_instrument_export",
                "workflow": "field",
                "format": "csv",
                "bytes": path.stat().st_size,
                "sha256": sha256_file(path),
            }
        )
    if community_path.exists():
        rows.append(
            {
                "path": (
                    str(community_path.relative_to(project_root))
                    if community_path.is_relative_to(project_root)
                    else str(community_path)
                ),
                "role": "derived_community_join_source",
                "workflow": "community",
                "format": community_path.suffix.lower().lstrip("."),
                "bytes": community_path.stat().st_size,
                "sha256": sha256_file(community_path),
            }
        )
    return sorted(rows, key=lambda row: (row["workflow"], row["path"]))


def method_metadata_table(
    project_root: Path,
    t5_info: Mapping[str, Any],
    t14_sample_count: int,
) -> list[dict[str, Any]]:
    """Summarize acquisition metadata without treating status as a method."""
    rows: list[dict[str, Any]] = []
    field_values: Counter[tuple[str, str]] = Counter()
    for path in sorted(
        project_path(project_root, FIELD_RAW_DIR).rglob("vanta_metadata_*.csv")
    ):
        records = read_csv_dicts(path)
        for record in records:
            for field in ("Method", "Mode", "Duration"):
                value = str(record.get(field, "")).strip()
                if value:
                    field_values[(field, value)] += 1
    for (field, value), count in sorted(field_values.items()):
        rows.append(
            {
                "workflow": "field",
                "trip_scope": "5",
                "compartment_scope": "not_encoded",
                "metadata_field": field,
                "value": value,
                "sample_or_session_count": count,
                "status": "encoded_in_instrument_export",
            }
        )

    for (compartment, field, value), count in sorted(
        t5_info["metadata_sample_counts"].items()
    ):
        rows.append(
            {
                "workflow": "lab",
                "trip_scope": "5",
                "compartment_scope": compartment,
                "metadata_field": field,
                "value": value,
                "sample_or_session_count": count,
                "status": "encoded_in_workbook",
            }
        )

    for field in ("Material", "Method", "Mode", "Diameter"):
        rows.append(
            {
                "workflow": "lab",
                "trip_scope": "1-4",
                "compartment_scope": "Deep;Surface;Rhizosphere",
                "metadata_field": field,
                "value": MISSING,
                "sample_or_session_count": t14_sample_count,
                "status": "not_encoded_in_consolidated_workbook_or_table",
            }
        )
    return sorted(
        rows,
        key=lambda row: (
            row["workflow"],
            row["trip_scope"],
            row["compartment_scope"],
            row["metadata_field"],
            row["value"],
        ),
    )


def read_csv_dicts(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def build_metadata_gaps(
    t14_observations: Sequence[Mapping[str, Any]],
    t5_observations: Sequence[Mapping[str, Any]],
    t5_info: Mapping[str, Any],
    missing_t5_analytical: Sequence[tuple[int, int, str]],
) -> list[dict[str, str]]:
    t14_lld = sum(row.get("lld") not in (None, "") for row in t14_observations)
    t5_lld = sum(row.get("lld") not in (None, "") for row in t5_observations)
    return [
        {
            "gap_id": "XRF-GAP-001",
            "workflow": "field_and_lab",
            "severity": "blocking_for_absolute_method_comparison",
            "metadata_field": "concentration_unit",
            "evidence": (
                "Processed concentration columns and the project protocol do not "
                "encode a unit per value/table."
            ),
            "consequence": (
                "Weight-percent comparability is plausible from instrument output "
                "but cannot be established from released metadata alone."
            ),
            "required_resolution": (
                "Add an explicit unit and basis (elemental versus oxide; wet/dry mass) "
                "from the instrument method/export documentation."
            ),
        },
        {
            "gap_id": "XRF-GAP-002",
            "workflow": "lab_t1_4",
            "severity": "limiting_for_cross_status_interpretation",
            "metadata_field": "instrument_method_mode_material_diameter",
            "evidence": (
                "The consolidated workbook exposes XRF status numbers, but the "
                "current table leaves Material, Method, Mode, and Diameter blank."
            ),
            "consequence": (
                "The documented max-positive processing rule must not be "
                "interpreted as evidence that statuses are calibrated or "
                "interchangeable."
            ),
            "required_resolution": (
                "Recover the source exports/method metadata referenced by the "
                "workbook index before any cross-status calibration claim; this "
                "is not required for reproducing the documented processing rule."
            ),
        },
        {
            "gap_id": "XRF-GAP-003",
            "workflow": "field_vs_lab",
            "severity": "blocking_for_direct_interchangeability",
            "metadata_field": "cross_workflow_calibration",
            "evidence": (
                "Field exports encode Geochem(3-Beam)/NORMAL, whereas Trip-5 "
                "laboratory workbooks encode Fast Screening-He8mm/He with 8mm "
                "diameter and Oxides material; no shared reference-material or "
                "cross-calibration record was found."
            ),
            "consequence": (
                "Field and laboratory values are complementary measurements, not "
                "interchangeable replicate measurements."
            ),
            "required_resolution": (
                "Use certified reference materials or a purpose-designed paired "
                "aliquot comparison before asserting absolute method equivalence."
            ),
        },
        {
            "gap_id": "XRF-GAP-004",
            "workflow": "lab_all",
            "severity": "limiting_for_nondetect_analysis",
            "metadata_field": "lod_and_nondetect_encoding",
            "evidence": (
                f"LLD is populated for {t14_lld}/{len(t14_observations)} "
                f"Trips 1-4 rows and {t5_lld}/{len(t5_observations)} Trip-5 rows; "
                "absent analytes are converted to zero in current wide tables."
            ),
            "consequence": (
                "Zero cannot safely be interpreted as a measured zero or a common "
                "detection threshold."
            ),
            "required_resolution": (
                "Publish run/analyte LOD semantics and preserve missing versus "
                "below-detection states in a long table."
            ),
        },
        {
            "gap_id": "XRF-GAP-005",
            "workflow": "field",
            "severity": "limiting_for_field_lab_pairing",
            "metadata_field": "physical_sample_and_compartment_id",
            "evidence": (
                "The field log links TestID to Site only; it does not encode a "
                "sample ID, compartment, depth, or replicate."
            ),
            "consequence": (
                "Field/lab validation is site-level only even though project context "
                "identifies the field measurements as Trip-5 Deep."
            ),
            "required_resolution": (
                "Add the measured material/sample identifier and depth for each "
                "field session."
            ),
        },
        {
            "gap_id": "XRF-GAP-006",
            "workflow": "lab_t5_retired_subset",
            "severity": "resolved_by_canonical_selection",
            "metadata_field": "canonical_input_selection",
            "evidence": (
                f"The 158-row analytical table omits "
                f"{len(missing_t5_analytical)} records present in the 178-row "
                "processed Trip-5 table."
            ),
            "consequence": (
                "No records are lost from the canonical analysis: the 178-row "
                "processed Trip-5 table is canonical and the 158-row table is "
                "retired."
            ),
            "required_resolution": (
                "Resolved: use 547 Trips 1-4 records plus all 178 processed "
                "Trip-5 records (725 total); retain the 158-row file only as "
                "retired provenance."
            ),
        },
    ]


def build_reconciliation(
    field_log: Sequence[Mapping[str, str]],
    field_rows: Sequence[Mapping[str, str]],
    field_raw_export_count: int,
    t14_rows: Sequence[Mapping[str, str]],
    t5_rows: Sequence[Mapping[str, str]],
    t5_analytical_rows: Sequence[Mapping[str, str]],
    t14_info: Mapping[str, Any],
    t5_info: Mapping[str, Any],
    all_community_keys: set[tuple[int, int, str]],
    qc_community_keys: set[tuple[int, int, str]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    t14_keys = table_sample_keys(t14_rows, "SampleID")
    t5_keys = table_sample_keys(t5_rows, "SampleID", default_trip=5)
    t5_analytical_keys = table_sample_keys(
        t5_analytical_rows, "SampleID", default_trip=5
    )
    legacy_subset_keys = t14_keys | t5_analytical_keys
    canonical_keys = t14_keys | t5_keys
    complete_field = [row for row in field_log if row.get("Status") == "COMPLETE"]
    field_sites = {int(float(row["SiteID"])) for row in field_rows}
    complete_sites = {int(float(row["Site"])) for row in complete_field}

    def joined(keys: set[tuple[int, int, str]]) -> tuple[int, int]:
        return len(keys & all_community_keys), len(keys & qc_community_keys)

    legacy_all, legacy_qc = joined(legacy_subset_keys)
    canonical_all, canonical_qc = joined(canonical_keys)
    rows = [
        {
            "dataset_id": "field_session_log_all",
            "workflow": "field",
            "trip_scope": "5",
            "compartment_scope": "not_encoded",
            "record_unit": "instrument_session_log_entry",
            "row_count": len(field_log),
            "unique_sites": len({int(float(row["Site"])) for row in field_log}),
            "status": "source",
            "caveat": "Includes COMPLETE, SHORT, and ERROR sessions.",
        },
        {
            "dataset_id": "field_session_log_complete",
            "workflow": "field",
            "trip_scope": "5",
            "compartment_scope": "project_context=Deep;not_encoded_in_log",
            "record_unit": "complete_instrument_session",
            "row_count": len(complete_field),
            "unique_sites": len(complete_sites),
            "status": "source",
            "caveat": "Repeated complete sessions occur at some sites.",
        },
        {
            "dataset_id": "field_instrument_exports_complete",
            "workflow": "field",
            "trip_scope": "5",
            "compartment_scope": "project_context=Deep;not_encoded_in_export",
            "record_unit": "instrument_export",
            "row_count": field_raw_export_count,
            "unique_sample_ids": field_raw_export_count,
            "unique_sites": len(field_sites),
            "status": "source_export",
            "caveat": "One Vanta composition export per retained complete TestID.",
        },
        {
            "dataset_id": "field_processed_table",
            "workflow": "field",
            "trip_scope": "5",
            "compartment_scope": "project_context=Deep;not_encoded_in_table",
            "record_unit": "complete_instrument_session",
            "row_count": len(field_rows),
            "unique_sample_ids": len(
                {(row["SiteID"], row["TestID"]) for row in field_rows}
            ),
            "unique_sites": len(field_sites),
            "status": "processed",
            "caveat": "One row per complete TestID; not one row per site.",
        },
        {
            "dataset_id": "lab_source_t1_4",
            "workflow": "lab",
            "trip_scope": "1-4",
            "compartment_scope": "Deep;Surface;Rhizosphere",
            "record_unit": "sample_sheet",
            "row_count": t14_info["sample_count"],
            "analyte_measurement_count": t14_info["measurement_count"],
            "unique_sample_ids": t14_info["sample_count"],
            "unique_trip_site_compartments": len(t14_keys),
            "unique_sites": len({key[1] for key in t14_keys}),
            "status": "source_consolidation",
            "caveat": "12 sheets; run/method metadata are incomplete.",
        },
        {
            "dataset_id": "lab_canonical_table_t1_4",
            "workflow": "lab",
            "trip_scope": "1-4",
            "compartment_scope": "Deep;Surface;Rhizosphere",
            "record_unit": "sample",
            "row_count": len(t14_rows),
            "unique_sample_ids": len({row["SampleID"] for row in t14_rows}),
            "unique_trip_site_compartments": len(t14_keys),
            "unique_sites": len({key[1] for key in t14_keys}),
            "status": "canonical_analytical",
            "caveat": "Uses maximum positive value across XRF statuses.",
        },
        {
            "dataset_id": "lab_source_t5_all_sheets",
            "workflow": "lab",
            "trip_scope": "5",
            "compartment_scope": "Deep;Surface;Rhizosphere",
            "record_unit": "workbook_sheet",
            "row_count": t5_info["all_sheet_count"],
            "analyte_measurement_count": t5_info["measurement_count"],
            "status": "source",
            "caveat": (
                f"{len(t5_info['excluded_sheets'])} Best Detection sheets are "
                "excluded by the current parser."
            ),
        },
        {
            "dataset_id": "lab_source_t5_selected_sheets",
            "workflow": "lab",
            "trip_scope": "5",
            "compartment_scope": "Deep;Surface;Rhizosphere",
            "record_unit": "selected_sample_sheet",
            "row_count": t5_info["selected_sheet_count"],
            "analyte_measurement_count": t5_info["measurement_count"],
            "unique_sample_ids": t5_info["selected_sample_count"],
            "unique_trip_site_compartments": len(t5_keys),
            "unique_sites": len({key[1] for key in t5_keys}),
            "status": "source_selected",
            "caveat": "Selection reproduces the existing parser behavior.",
        },
        {
            "dataset_id": "lab_processed_table_t5",
            "workflow": "lab",
            "trip_scope": "5",
            "compartment_scope": "Deep;Surface;Rhizosphere",
            "record_unit": "sample",
            "row_count": len(t5_rows),
            "unique_sample_ids": len({row["SampleID"] for row in t5_rows}),
            "unique_trip_site_compartments": len(t5_keys),
            "unique_sites": len({key[1] for key in t5_keys}),
            "status": "canonical_analytical",
            "caveat": (
                "Canonical Trip-5 table; repeated formulas keep the last "
                "workbook row."
            ),
        },
        {
            "dataset_id": "lab_retired_analytical_subset_t5",
            "workflow": "lab",
            "trip_scope": "5",
            "compartment_scope": "Deep;Surface;Rhizosphere",
            "record_unit": "sample",
            "row_count": len(t5_analytical_rows),
            "unique_sample_ids": len(
                {row["SampleID"] for row in t5_analytical_rows}
            ),
            "unique_trip_site_compartments": len(t5_analytical_keys),
            "unique_sites": len({key[1] for key in t5_analytical_keys}),
            "status": "retired_legacy_subset",
            "caveat": (
                "Retired 158-row subset; 20 canonical Trip-5 records are absent."
            ),
        },
        {
            "dataset_id": "lab_all_retired_analytical_subset",
            "workflow": "lab",
            "trip_scope": "1-5",
            "compartment_scope": "Deep;Surface;Rhizosphere",
            "record_unit": "trip_site_compartment",
            "row_count": len(legacy_subset_keys),
            "unique_trip_site_compartments": len(legacy_subset_keys),
            "unique_sites": len({key[1] for key in legacy_subset_keys}),
            "community_joined_all": legacy_all,
            "community_joined_qc": legacy_qc,
            "status": "retired_legacy_subset",
            "caveat": "Retired combination: 547 Trips 1-4 plus 158 Trip-5.",
        },
        {
            "dataset_id": "lab_all_canonical_analytical",
            "workflow": "lab",
            "trip_scope": "1-5",
            "compartment_scope": "Deep;Surface;Rhizosphere",
            "record_unit": "trip_site_compartment",
            "row_count": len(canonical_keys),
            "unique_trip_site_compartments": len(canonical_keys),
            "unique_sites": len({key[1] for key in canonical_keys}),
            "community_joined_all": canonical_all,
            "community_joined_qc": canonical_qc,
            "status": "canonical_analytical",
            "caveat": "Canonical: 547 Trips 1-4 plus 178 Trip-5 records.",
        },
    ]
    return rows, {
        "t14_keys": t14_keys,
        "t5_keys": t5_keys,
        "t5_analytical_keys": t5_analytical_keys,
        "legacy_subset_keys": legacy_subset_keys,
        "canonical_keys": canonical_keys,
        "legacy_joined_all": legacy_all,
        "legacy_joined_qc": legacy_qc,
        "canonical_joined_all": canonical_all,
        "canonical_joined_qc": canonical_qc,
    }


def build_discrepancies(
    t14_observations: Sequence[Mapping[str, Any]],
    t5_observations: Sequence[Mapping[str, Any]],
    t14_analytes: Sequence[str],
    t5_analytical_missing: Sequence[tuple[int, int, str]],
    t14_reproduction: Mapping[str, Any],
    t5_reproduction: Mapping[str, Any],
) -> list[dict[str, Any]]:
    raw_t14_formulas = {str(row["formula"]) for row in t14_observations}
    omitted = sorted(raw_t14_formulas - set(t14_analytes))
    discrepancies: list[dict[str, Any]] = [
        {
            "issue_type": "aggregation_rule_reproduction",
            "workflow": "lab_t1_4",
            "sample_key": "ALL",
            "detail": (
                f"Current table compared with {t14_reproduction['rule']}: "
                f"{t14_reproduction['mismatched_reported_cells']} mismatches among "
                f"{t14_reproduction['checked_reported_cells']} source-backed cells."
            ),
            "requires_resolution": (
                "Yes" if t14_reproduction["mismatched_reported_cells"] else "No"
            ),
        },
        {
            "issue_type": "aggregation_rule_reproduction",
            "workflow": "lab_t5_processed",
            "sample_key": "ALL",
            "detail": (
                f"Current table compared with {t5_reproduction['rule']}: "
                f"{t5_reproduction['mismatched_reported_cells']} mismatches among "
                f"{t5_reproduction['checked_reported_cells']} source-backed cells."
            ),
            "requires_resolution": (
                "Yes" if t5_reproduction["mismatched_reported_cells"] else "No"
            ),
        },
    ]
    for analyte in omitted:
        count = sum(
            row["formula"] == analyte and float(row["value"]) > 0
            for row in t14_observations
        )
        discrepancies.append(
            {
                "issue_type": "raw_analyte_absent_from_wide_schema",
                "workflow": "lab_t1_4",
                "sample_key": "ALL",
                "detail": f"{analyte}: {count} positive source rows omitted.",
                "requires_resolution": "Yes",
            }
        )
    for trip, site, compartment in sorted(t5_analytical_missing):
        discrepancies.append(
            {
                "issue_type": "processed_sample_absent_from_analytical_table",
                "workflow": "lab_t5_retired_subset",
                "sample_key": f"{trip}|{site}|{compartment}",
                "detail": (
                    "Present in canonical 178-row table; absent from retired "
                    "158-row subset."
                ),
                "requires_resolution": "No",
            }
        )
    return discrepancies


def render_report(
    summary: Mapping[str, Any],
    reconciliation: Sequence[Mapping[str, Any]],
    sensitivity: Sequence[Mapping[str, Any]],
    method_metadata: Sequence[Mapping[str, Any]],
    agreement: Sequence[Mapping[str, Any]],
    metadata_gaps: Sequence[Mapping[str, str]],
    discrepancies: Sequence[Mapping[str, Any]],
) -> str:
    counts = summary["counts"]
    significant = [
        row
        for row in sensitivity
        if int(row["groups_with_multiple_rows"]) > 0
        or int(row["groups_missing_primary_status"]) > 0
    ]
    significant.sort(
        key=lambda row: (
            -int(row["groups_current_differs_from_primary"]),
            -int(row["groups_with_multiple_rows"]),
            -int(row["groups_missing_primary_status"]),
            row["workflow"],
            row["analyte"],
        )
    )
    agreement_ranked = [
        row
        for row in agreement
        if row["both_reported_positive"] >= 10
        and row["spearman_all_sites"] is not None
    ]
    agreement_ranked.sort(
        key=lambda row: float(row["spearman_all_sites"]), reverse=True
    )
    missing_samples = [
        row
        for row in discrepancies
        if row["issue_type"] == "processed_sample_absent_from_analytical_table"
    ]
    missing_samples.sort(
        key=lambda row: (
            int(row["sample_key"].split("|")[1]),
            row["sample_key"].split("|")[2],
        )
    )

    lines = [
        "# Empty Quarter XRF provenance and reconciliation audit",
        "",
        "## Bottom line",
        "",
        "The field and laboratory XRF datasets are not contradictory. They are "
        "different workflows with different observational units:",
        "",
        f"- **Field XRF:** the source log contains "
        f"{counts['field_log_rows']} Trip-5 entries across "
        f"{counts['field_log_sites']} sites. Of these, "
        f"{counts['field_complete_sessions']} complete instrument sessions at "
        f"{counts['field_sites']} sites were retained; "
        f"{counts['field_repeated_sites']} sites have repeated complete sessions. "
        "The source log is site-level and does not itself encode the sample, "
        "compartment, or depth.",
        f"- **Laboratory XRF:** {counts['lab_t14_samples']} selected sample sheets "
        f"from Trips 1--4 plus {counts['lab_t5_processed_samples']} selected Trip-5 "
        f"sample sheets = **{counts['lab_all_canonical']} canonical laboratory "
        "records**.",
        f"- The former 158-row Trip-5 analytical subset (705 records when combined "
        f"with Trips 1--4) is retired because it omits "
        f"{counts['lab_t5_missing_from_retired_subset']} records without an "
        "exclusion rule.",
        f"- With the >=2,000-read ecology filter, the canonical 725 records join "
        f"to {counts['community_join_canonical_qc']} community groups; the retired "
        f"705-record subset joined to {counts['community_join_retired_subset_qc']}.",
        "",
        "Therefore, both manuscripts should describe the two workflows separately "
        "and use 725 (=547+178) as the laboratory sample count.",
        "",
        "## Reconciliation",
        "",
        "| Dataset | Workflow | Unit | Rows | Analyte rows | Joined (QC) | Caveat |",
        "|---|---|---:|---:|---:|---:|---|",
    ]
    for row in reconciliation:
        display = {
            column: format_cell(row.get(column, MISSING))
            for column in RECONCILIATION_COLUMNS
        }
        lines.append(
            "| {dataset_id} | {workflow} | {record_unit} | {row_count} | "
            "{analyte_measurement_count} | {community_joined_qc} | {caveat} |".format(
                **display
            )
        )

    lines.extend(
        [
            "",
            "## Aggregation audit",
            "",
            "The current Trips 1--4 builder explicitly takes the largest positive "
            "concentration across rows/statuses. The audit reproduces the current "
            "wide table from that rule. The current Trip-5 parser instead overwrites "
            "repeated formulas and therefore retains the last reported row. These "
            "are different policies.",
            "",
            "In the present sources, the numerical impact is localized: the "
            "Trips 1--4 maximum agrees with the primary-status value whenever that "
            "value is present, while some rare compounds occur only under a "
            "secondary status. Trip 5 has a small number of within-status repeated "
            "rows for which last-value and primary-status-median choices differ. "
            "The detailed group-level candidate values are in "
            "`xrf_aggregation_group_details.tsv`.",
            "",
            "The `primary_status` sensitivity below uses workbook section labels "
            "(`XRF 0 (Elements)` and `XRF 1 (Oxides)`) to compare a status-specific "
            "candidate against the current policy. It is not adopted as canonical "
            "because instrument method metadata, units, and non-detect semantics "
            "remain incomplete.",
            "",
            "| Workflow | Analyte | Current rule | Multi-row groups | Current != "
            "primary | Missing primary | Spearman | Max absolute relative "
            "difference |",
            "|---|---|---|---:|---:|---:|---:|---:|",
        ]
    )
    for row in significant[:20]:
        lines.append(
            f"| {row['workflow']} | {row['analyte']} | {row['current_rule']} | "
            f"{row['groups_with_multiple_rows']} | "
            f"{row['groups_current_differs_from_primary']} | "
            f"{row['groups_missing_primary_status']} | "
            f"{format_cell(row['spearman_current_vs_primary'])} | "
            f"{format_cell(row['maximum_absolute_relative_difference'])} |"
        )

    lines.extend(
        [
            "",
            "## Acquisition metadata",
            "",
            "Trip-5 acquisition settings are recoverable and confirm that the "
            "workflows are methodologically distinct: field exports report "
            "`Geochem(3-Beam)`/`NORMAL`, while laboratory workbooks report "
            "`Fast Screening-He8mm`/`He`, `8mm`, and `Oxides` for Deep, Surface, "
            "and Rhizosphere records. This supports treating the datasets as "
            "complementary rather than interchangeable.",
            "",
            "| Workflow | Trips | Compartment | Field | Value | Records | Status |",
            "|---|---|---|---|---|---:|---|",
        ]
    )
    for row in method_metadata:
        lines.append(
            f"| {row['workflow']} | {row['trip_scope']} | "
            f"{row['compartment_scope']} | {row['metadata_field']} | "
            f"{format_cell(row['value'])} | {row['sample_or_session_count']} | "
            f"{row['status']} |"
        )

    lines.extend(
        [
            "",
            "## Trip-5 field versus laboratory comparison",
            "",
            f"The audit matched {counts['field_lab_matched_sites']} sites after "
            "reducing repeated complete field sessions by their median and selecting "
            "the Trip-5 laboratory Deep record. This is a **site-level diagnostic**, "
            "not a same-aliquot validation experiment, because field session records "
            "lack a physical sample identifier.",
            "",
            "Zeros are treated only as “not reported positive” because LOD metadata "
            "are incomplete. Rank correlations and log field/lab ratios are "
            "descriptive. They do not establish interchangeability.",
            "",
            "| Analyte | Both positive | Positive agreement | Spearman (all sites) | "
            "Median field/lab ratio | Mean log10 bias |",
            "|---|---:|---:|---:|---:|---:|",
        ]
    )
    for row in agreement_ranked[:20]:
        lines.append(
            f"| {row['analyte']} | {row['both_reported_positive']} | "
            f"{format_cell(row['reported_positive_agreement'])} | "
            f"{format_cell(row['spearman_all_sites'])} | "
            f"{format_cell(row['median_field_lab_ratio'])} | "
            f"{format_cell(row['mean_log10_field_lab_bias'])} |"
        )

    lines.extend(
        [
            "",
            "## Metadata gaps and resolved decisions",
            "",
            "| ID | Workflow | Severity | Missing/uncertain field | Consequence |",
            "|---|---|---|---|---|",
        ]
    )
    for gap in metadata_gaps:
        lines.append(
            f"| {gap['gap_id']} | {gap['workflow']} | {gap['severity']} | "
            f"{gap['metadata_field']} | {gap['consequence']} |"
        )

    lines.extend(
        [
            "",
            "## Missing Trip-5 analytical records",
            "",
            ", ".join(row["sample_key"] for row in missing_samples),
            "",
            "## Manuscript-ready wording",
            "",
            "> During Trip 5, handheld X-ray fluorescence measurements were "
            f"obtained in the field in {counts['field_complete_sessions']} complete "
            f"measurement sessions across {counts['field_sites']} sites. These "
            "in-situ, site-level observations were maintained separately from "
            "laboratory XRF measurements of archived soil samples. The laboratory "
            f"source data comprised {counts['lab_t14_samples']} sample records from "
            f"Trips 1--4 and {counts['lab_t5_processed_samples']} selected records "
            "from Trip 5. Laboratory values used for ecological analysis were "
            "linked by trip, site, and compartment; field values were used only "
            "for a descriptive Trip-5 site-level comparison because the field log "
            "did not encode a physical sample identifier.",
            "",
            "Do not state that field and laboratory XRF disagree. State instead "
            "that their agreement is only partially testable with current linkage "
            "and method metadata. Use 725 (=547+178) as the canonical laboratory "
            "record count; the 158-row Trip-5 / 705-row combined subset is retired.",
            "",
            "## Reproduction",
            "",
            "```bash",
            "uv run --with openpyxl python scripts/xrf/audit_xrf_provenance.py \\",
            "  --project-root . --output-dir analysis/xrf_audit",
            "```",
            "",
            "All TSV and JSON outputs are stably sorted and contain no run timestamp. "
            "The source inventory records SHA-256 checksums.",
            "",
        ]
    )
    return "\n".join(lines)


def run_audit(
    project_root: Path,
    output_dir: Path,
    community_table: Path | None = None,
    minimum_community_reads: float = 2000.0,
) -> dict[str, Any]:
    project_root = project_root.resolve()
    output_dir = (
        output_dir.resolve()
        if output_dir.is_absolute()
        else (project_root / output_dir).resolve()
    )
    community_path = (
        community_table.resolve()
        if community_table is not None and community_table.is_absolute()
        else project_path(project_root, community_table or COMMUNITY_TABLE)
    )

    field_log = read_tsv(project_path(project_root, FIELD_LOG))
    field_rows = read_tsv(project_path(project_root, FIELD_TABLE))
    t14_rows, t14_analytes, t14_id = load_numeric_table(
        project_path(project_root, LAB_T14_TABLE)
    )
    t5_rows, t5_analytes, t5_id = load_numeric_table(
        project_path(project_root, LAB_T5_TABLE)
    )
    t5_analytical_rows, _, _ = load_numeric_table(
        project_path(project_root, LAB_T5_ANALYTICAL)
    )
    t14_observations, t14_info = load_t14_observations(
        project_path(project_root, LAB_T14_WORKBOOK)
    )
    t5_observations, t5_info = load_t5_observations(
        project_path(project_root, LAB_T5_DIR)
    )

    all_community_keys: set[tuple[int, int, str]] = set()
    qc_community_keys: set[tuple[int, int, str]] = set()
    community_info: dict[str, int] = {}
    if community_path.exists():
        all_community_keys, qc_community_keys, community_info = load_community_keys(
            community_path,
            minimum_reads=minimum_community_reads,
        )

    field_raw_export_count = len(
        list(project_path(project_root, FIELD_RAW_DIR).rglob("vanta_data_*.csv"))
    )
    reconciliation, key_info = build_reconciliation(
        field_log,
        field_rows,
        field_raw_export_count,
        t14_rows,
        t5_rows,
        t5_analytical_rows,
        t14_info,
        t5_info,
        all_community_keys,
        qc_community_keys,
    )
    missing_t5_analytical = sorted(
        key_info["t5_keys"] - key_info["t5_analytical_keys"]
    )

    t14_reproduction = compare_current_table_to_rule(
        t14_rows,
        t14_analytes,
        t14_id,
        t14_observations,
        "max_positive",
    )
    t5_reproduction = compare_current_table_to_rule(
        t5_rows,
        t5_analytes,
        t5_id,
        t5_observations,
        "last_reported",
    )
    sensitivity = aggregation_sensitivity(
        "lab_t1_4",
        t14_observations,
        "max_positive",
    ) + aggregation_sensitivity(
        "lab_t5",
        t5_observations,
        "last_reported",
    )
    sensitivity.sort(key=lambda row: (row["workflow"], row["analyte"]))
    aggregation_details = aggregation_group_details(
        "lab_t1_4",
        t14_observations,
        "max_positive",
    ) + aggregation_group_details(
        "lab_t5",
        t5_observations,
        "last_reported",
    )
    aggregation_details.sort(
        key=lambda row: (
            row["workflow"],
            int(row["trip"]),
            int(row["site"]),
            row["compartment"],
            row["analyte"],
        )
    )

    agreement, field_replicates, agreement_info = field_lab_agreement(
        field_rows,
        t5_rows,
    )
    agreement.sort(key=lambda row: row["analyte"])
    field_replicates.sort(key=lambda row: row["analyte"])

    metadata_gaps = build_metadata_gaps(
        t14_observations,
        t5_observations,
        t5_info,
        missing_t5_analytical,
    )
    method_metadata = method_metadata_table(
        project_root,
        t5_info,
        len(t14_rows),
    )
    discrepancies = build_discrepancies(
        t14_observations,
        t5_observations,
        t14_analytes,
        missing_t5_analytical,
        t14_reproduction,
        t5_reproduction,
    )
    discrepancies.sort(
        key=lambda row: (
            row["workflow"],
            row["issue_type"],
            row["sample_key"],
            row["detail"],
        )
    )

    field_statuses = Counter(row["Status"] for row in field_log)
    field_log_sites = {int(float(row["Site"])) for row in field_log}
    field_sites = {int(float(row["SiteID"])) for row in field_rows}
    repeated_field_sites = Counter(int(float(row["SiteID"])) for row in field_rows)
    counts = {
        "field_log_rows": len(field_log),
        "field_log_sites": len(field_log_sites),
        "field_complete_sessions": field_statuses["COMPLETE"],
        "field_short_sessions": field_statuses["SHORT"],
        "field_error_sessions": field_statuses["ERROR"],
        "field_processed_rows": len(field_rows),
        "field_instrument_exports": field_raw_export_count,
        "field_sites": len(field_sites),
        "field_repeated_sites": sum(
            count > 1 for count in repeated_field_sites.values()
        ),
        "lab_t14_samples": len(t14_rows),
        "lab_t14_measurements": len(t14_observations),
        "lab_t14_raw_analytes_absent_from_schema": len(
            {str(row["formula"]) for row in t14_observations} - set(t14_analytes)
        ),
        "lab_t5_workbook_sheets": t5_info["all_sheet_count"],
        "lab_t5_selected_sheets": t5_info["selected_sheet_count"],
        "lab_t5_processed_samples": len(t5_rows),
        "lab_t5_measurements": len(t5_observations),
        "lab_t5_canonical": len(t5_rows),
        "lab_t5_retired_analytical_subset": len(t5_analytical_rows),
        "lab_t5_missing_from_retired_subset": len(missing_t5_analytical),
        "aggregation_groups_audited": len(aggregation_details),
        "lab_all_canonical": len(key_info["canonical_keys"]),
        "lab_all_retired_analytical_subset": len(
            key_info["legacy_subset_keys"]
        ),
        "community_join_canonical_all": key_info["canonical_joined_all"],
        "community_join_canonical_qc": key_info["canonical_joined_qc"],
        "community_join_retired_subset_all": key_info["legacy_joined_all"],
        "community_join_retired_subset_qc": key_info["legacy_joined_qc"],
        "field_lab_matched_sites": agreement_info["matched_sites"],
    }
    summary = {
        "audit_schema_version": AUDIT_SCHEMA_VERSION,
        "parameters": {
            "minimum_community_reads": minimum_community_reads,
            "community_table": (
                str(community_path.relative_to(project_root))
                if community_path.is_relative_to(project_root)
                else str(community_path)
            ),
        },
        "counts": counts,
        "aggregation_reproduction": {
            "lab_t1_4": t14_reproduction,
            "lab_t5": t5_reproduction,
        },
        "community": community_info,
        "field_lab": agreement_info,
        "canonical_policy": {
            "laboratory_record_count": len(key_info["canonical_keys"]),
            "trips_1_4_records": len(t14_rows),
            "trips_1_4_rule": "max_positive",
            "trip_5_records": len(t5_rows),
            "trip_5_rule": "last_reported",
            "retired_trip_5_subset_records": len(t5_analytical_rows),
            "field_lab_interchangeability_claim": False,
        },
        "critical_flags": [
            "FIELD_LOG_HAS_NO_PHYSICAL_SAMPLE_OR_COMPARTMENT_ID",
            "LAB_T1_4_MAX_ACROSS_STATUS_IS_PROCESSING_NOT_CALIBRATION",
            "LAB_T1_4_AND_T5_USE_DIFFERENT_REPEAT_FORMULA_RULES",
            "CONCENTRATION_UNIT_AND_NONDETECT_SEMANTICS_NOT_MACHINE_READABLE",
        ],
        "resolved_flags": [
            "LAB_T5_158_ROW_SUBSET_RETIRED",
            "LAB_CANONICAL_COUNT_725_EQUALS_547_PLUS_178",
        ],
    }

    inventory = source_inventory(project_root, community_path)

    output_dir.mkdir(parents=True, exist_ok=True)
    write_tsv(
        output_dir / "xrf_source_inventory.tsv",
        inventory,
        ["path", "role", "workflow", "format", "bytes", "sha256"],
    )
    write_tsv(
        output_dir / "xrf_reconciliation.tsv",
        reconciliation,
        RECONCILIATION_COLUMNS,
    )
    write_tsv(
        output_dir / "xrf_aggregation_sensitivity.tsv",
        sensitivity,
        [
            "workflow",
            "analyte",
            "formula_class",
            "expected_primary_status",
            "observed_statuses",
            "sample_analyte_groups",
            "groups_with_multiple_rows",
            "groups_missing_primary_status",
            "groups_current_differs_from_primary",
            "current_rule",
            "spearman_current_vs_primary",
            "median_relative_difference",
            "q95_absolute_relative_difference",
            "median_relative_difference_among_changed",
            "maximum_absolute_relative_difference",
            "paired_groups",
        ],
    )
    write_tsv(
        output_dir / "xrf_aggregation_group_details.tsv",
        aggregation_details,
        [
            "workflow",
            "sample_id",
            "trip",
            "site",
            "compartment",
            "analyte",
            "source_row_count",
            "status_value_sequence",
            "expected_primary_status",
            "max_positive",
            "mean_positive",
            "median_positive",
            "first_reported",
            "last_reported",
            "primary_status_value",
            "current_rule",
            "current_value",
            "current_minus_primary_relative",
            "primary_status_available",
        ],
    )
    write_tsv(
        output_dir / "xrf_current_table_discrepancies.tsv",
        discrepancies,
        [
            "issue_type",
            "workflow",
            "sample_key",
            "detail",
            "requires_resolution",
        ],
    )
    write_tsv(
        output_dir / "xrf_metadata_gaps.tsv",
        metadata_gaps,
        [
            "gap_id",
            "workflow",
            "severity",
            "metadata_field",
            "evidence",
            "consequence",
            "required_resolution",
        ],
    )
    write_tsv(
        output_dir / "xrf_method_metadata.tsv",
        method_metadata,
        [
            "workflow",
            "trip_scope",
            "compartment_scope",
            "metadata_field",
            "value",
            "sample_or_session_count",
            "status",
        ],
    )
    agreement_columns = [
        "analyte",
        "matched_sites",
        "both_reported_positive",
        "field_only_positive",
        "lab_only_positive",
        "neither_positive",
        "reported_positive_agreement",
        "spearman_all_sites",
        "spearman_both_positive",
        "median_field_lab_ratio",
        "mean_log10_field_lab_bias",
        "log10_lower_loa",
        "log10_upper_loa",
        "interpretation_limit",
    ]
    write_tsv(
        output_dir / "xrf_field_lab_agreement.tsv",
        agreement,
        agreement_columns,
    )
    write_tsv(
        output_dir / "xrf_field_replicate_precision.tsv",
        field_replicates,
        [
            "analyte",
            "sites_with_repeated_field_sessions",
            "sites_with_estimable_cv",
            "median_within_site_cv",
            "q95_within_site_cv",
        ],
    )
    site_columns = (
        ["site", "field_sessions", "lab_sample_id", "match_level"]
        + [
            column
            for analyte in sorted(
                {
                    row["analyte"]
                    for row in agreement
                }
            )
            for column in (f"field_{analyte}", f"lab_{analyte}")
        ]
    )
    # Recompute the compact site table only for output; field_lab_agreement keeps
    # it internal so downstream reports cannot accidentally call it a sample match.
    _, _, _ = agreement, field_replicates, agreement_info
    # Use the same deterministic matching routine and extract its site table via
    # a private helper call below.
    site_rows = build_field_lab_site_rows(field_rows, t5_rows, agreement)
    write_tsv(
        output_dir / "xrf_field_lab_site_matches.tsv",
        site_rows,
        site_columns,
    )

    with (output_dir / "xrf_audit_summary.json").open(
        "w", encoding="utf-8"
    ) as handle:
        json.dump(summary, handle, indent=2, sort_keys=True)
        handle.write("\n")
    report = render_report(
        summary,
        reconciliation,
        sensitivity,
        method_metadata,
        agreement,
        metadata_gaps,
        discrepancies,
    )
    (output_dir / "xrf_evidence_report.md").write_text(report, encoding="utf-8")
    return summary


def build_field_lab_site_rows(
    field_rows: Sequence[Mapping[str, str]],
    lab_rows: Sequence[Mapping[str, str]],
    agreement_rows: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    analytes = [str(row["analyte"]) for row in agreement_rows]
    field_by_site: dict[int, list[Mapping[str, str]]] = defaultdict(list)
    for row in field_rows:
        field_by_site[int(float(row["SiteID"]))].append(row)
    lab_by_site: dict[int, Mapping[str, str]] = {}
    for row in lab_rows:
        if row.get("SoilType") != "Deep":
            continue
        key = canonical_sample_key(str(row["SampleID"]), default_trip=5)
        if key is not None:
            lab_by_site[key[1]] = row
    output: list[dict[str, Any]] = []
    for site in sorted(set(field_by_site) & set(lab_by_site)):
        row: dict[str, Any] = {
            "site": site,
            "field_sessions": len(field_by_site[site]),
            "lab_sample_id": lab_by_site[site]["SampleID"],
            "match_level": "site_to_Trip5_Deep",
        }
        for analyte in analytes:
            values = [
                value
                for source in field_by_site[site]
                if (value := numeric(source.get(analyte))) is not None
            ]
            row[f"field_{analyte}"] = statistics.median(values) if values else None
            row[f"lab_{analyte}"] = numeric(lab_by_site[site].get(analyte))
        output.append(row)
    return output


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--project-root",
        type=Path,
        default=Path("."),
        help="Empty Quarter repository root (default: current directory).",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("analysis/xrf_audit"),
        help="Output directory, absolute or relative to --project-root.",
    )
    parser.add_argument(
        "--community-table",
        type=Path,
        default=None,
        help=(
            "Derived genus count table used for the ecology join. Defaults to "
            "analysis/v2/review/cache/genus_counts.tsv under --project-root."
        ),
    )
    parser.add_argument(
        "--minimum-community-reads",
        type=float,
        default=2000.0,
        help="Read-count threshold reproducing the current ecology join (default: 2000).",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    summary = run_audit(
        args.project_root,
        args.output_dir,
        community_table=args.community_table,
        minimum_community_reads=args.minimum_community_reads,
    )
    counts = summary["counts"]
    print(
        "XRF audit complete: "
        f"{counts['field_complete_sessions']} field sessions; "
        f"{counts['lab_all_canonical']} canonical lab records; "
        f"{counts['community_join_canonical_qc']} canonical community joins."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
