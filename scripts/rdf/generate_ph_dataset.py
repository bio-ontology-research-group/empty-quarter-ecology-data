#!/usr/bin/env python3
"""Audit a versioned soil-pH workbook and build a versioned RDF module.

This generator is intentionally separate from the released Rub al-Khali KG.
It accepts only source rows that pass the workbook's declared calibration and
quality-control rules, resolve to exactly one existing specimen, and carry an
Excel, ISO or day-first ``DD/MM/YYYY`` date that parses no later than
``--as-of``. It never repairs a source cell silently. Pending measurements,
depleted samples, unparseable or post-cutoff dates, and suspected column shifts
remain in the audit table but do not enter the RDF graph.

The RDF follows the project's four-individual SIO measurement pattern:

    specimen -> acidity quality -> pH measurement value <- measuring process

The value is both ``sio:SIO_000070`` (measurement value) and
``sio:SIO_001089`` (pH), uses ``uo:UO_0000196``, and carries its numeric value
through ``sio:SIO_000300``.  The operational method is represented as a
protocol because soil pH depends on the extraction medium and ratio.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import xml.etree.ElementTree as ET
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from rdflib import DCTERMS, OWL, RDF, RDFS, XSD, Graph, Literal, Namespace, URIRef
from rdflib.namespace import split_uri


BASE = Namespace("https://rubalkhali.science/kb/")
SIO = Namespace("http://semanticscience.org/resource/")
PATO = Namespace("http://purl.obolibrary.org/obo/PATO_")
UO = Namespace("http://purl.obolibrary.org/obo/UO_")
DC = Namespace("http://purl.org/dc/elements/1.1/")
DCAT = Namespace("http://www.w3.org/ns/dcat#")

SCHEMA_VERSION = "ph-measurement-v1"
PROTOCOL_KEY = "EQ-PH-CACL2-1G-v1.0.0"
ADMITTED = "ADMITTED_MEASUREMENT"
EXPECTED_HEADERS = [
    "Sample ID",
    "Site",
    "Compartment",
    "Rep",
    "Date",
    "Calibration",
    "Slope %",
    "pH 10 check",
    "pH reading",
    "Notes",
]
DECLARED_SAMPLE_COUNTS = {1: 330, 2: 27, 3: 450, 4: 177, 5: 234}
TRIP_SHEETS = {trip: f"Trip {trip}" for trip in range(1, 6)}
SAMPLE_RE = re.compile(
    r"^(?P<prefix>[TFSV])?(?P<site>\d+)(?P<compartment>PR|P|D|S)r(?P<rep>\d+)$"
)
PREFIX_TRIP = {"": 1, "T": 2, "F": 3, "S": 4, "V": 5}
COMPARTMENT = {
    "S": "Surface",
    "D": "Deep",
    "P": "Rhizosphere",
    "PR": "Rhizosphere",
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def stable_iri(role: str, key: str) -> tuple[URIRef, str]:
    canonical = f"{SCHEMA_VERSION}|{role}|{key}"
    digest = hashlib.sha256(canonical.encode("utf-8")).hexdigest()
    return URIRef(str(BASE) + f"RAK_PH_{role.upper()}_{digest}"), canonical


def is_number(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(float(value))


def clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_measurement_date(value: object, as_of: date) -> tuple[str, str]:
    if value is None or clean_text(value) == "":
        return "", "missing"
    parsed: date | None = None
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    else:
        text = clean_text(value)
        for pattern in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                parsed = datetime.strptime(text, pattern).date()
                break
            except ValueError:
                continue
    if parsed is None:
        return "", "unparseable"
    if parsed > as_of:
        return "", "future_or_day_month_ambiguous"
    return parsed.isoformat(), "accepted"


def parse_sample_id(sample_id: str) -> dict[str, Any] | None:
    match = SAMPLE_RE.fullmatch(sample_id.replace(" ", ""))
    if match is None:
        return None
    prefix = match.group("prefix") or ""
    return {
        "trip": PREFIX_TRIP[prefix],
        "site": int(match.group("site")),
        "compartment": COMPARTMENT[match.group("compartment")],
        "replicate": int(match.group("rep")),
    }


def read_sample_ledger(path: Path) -> dict[tuple[int, str], dict[str, str]]:
    result: dict[tuple[int, str], dict[str, str]] = {}
    with path.open(encoding="utf-8") as handle:
        for row in csv.DictReader(handle, delimiter="\t"):
            trip_text = row["trip"].replace("Trip", "").strip()
            if not trip_text.isdigit():
                continue
            key = (int(trip_text), row["sample_id"].strip())
            if key in result:
                raise ValueError(f"Duplicate sample-ledger key: {key}")
            result[key] = row
    return result


def read_sample_iris(path: Path) -> dict[str, URIRef]:
    graph = Graph().parse(path)
    candidates: dict[str, set[URIRef]] = {}
    for subject, identifier in graph.subject_objects(DC.identifier):
        candidates.setdefault(str(identifier).strip(), set()).add(URIRef(subject))
    duplicates = {key: values for key, values in candidates.items() if len(values) != 1}
    if duplicates:
        preview = ", ".join(sorted(duplicates)[:5])
        raise ValueError(f"Non-unique sample identifiers in sample ABox: {preview}")
    return {key: next(iter(values)) for key, values in candidates.items()}


def write_tsv(path: Path, columns: list[str], rows: Iterable[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: clean_text(row.get(column)) for column in columns})


def serialize_sorted_turtle(graph: Graph, path: Path) -> None:
    """Write deterministic N-Triples syntax, which is valid Turtle."""
    lines = [
        "# Versioned soil-pH graph; canonical import is controlled separately.",
        "# N-Triples syntax is used for deterministic byte serialization.",
    ]
    triples = sorted(
        graph,
        key=lambda triple: tuple(term.n3() for term in triple),
    )
    lines.extend(f"{s.n3()} {p.n3()} {o.n3()} ." for s, p, o in triples)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def serialize_rdfxml(graph: Graph, path: Path) -> None:
    """Write deterministic, flat RDF/XML for OWL-oriented consumers."""
    for prefix, namespace in (
        ("rdf", RDF),
        ("rdfs", RDFS),
        ("owl", OWL),
        ("dcterms", DCTERMS),
        ("sio", SIO),
    ):
        ET.register_namespace(prefix, str(namespace))

    root = ET.Element(ET.QName(str(RDF), "RDF"))
    for subject in sorted(set(graph.subjects()), key=lambda node: node.n3()):
        if not isinstance(subject, URIRef):
            raise ValueError(f"RDF/XML serializer does not support subject {subject!r}")
        description = ET.SubElement(root, ET.QName(str(RDF), "Description"))
        description.set(ET.QName(str(RDF), "about"), str(subject))
        statements = sorted(
            graph.predicate_objects(subject),
            key=lambda pair: (pair[0].n3(), pair[1].n3()),
        )
        for predicate, obj in statements:
            namespace, local = split_uri(predicate)
            element = ET.SubElement(description, ET.QName(str(namespace), local))
            if isinstance(obj, URIRef):
                element.set(ET.QName(str(RDF), "resource"), str(obj))
            elif isinstance(obj, Literal):
                if obj.datatype is not None:
                    element.set(ET.QName(str(RDF), "datatype"), str(obj.datatype))
                if obj.language is not None:
                    element.set(
                        ET.QName("http://www.w3.org/XML/1998/namespace", "lang"),
                        obj.language,
                    )
                element.text = str(obj)
            else:
                raise ValueError(f"RDF/XML serializer does not support object {obj!r}")
    ET.indent(root, space="  ")
    tree = ET.ElementTree(root)
    tree.write(path, encoding="utf-8", xml_declaration=True, short_empty_elements=True)
    with path.open("ab") as handle:
        handle.write(b"\n")


def audit_workbook(
    workbook_path: Path,
    sample_ledger: dict[tuple[int, str], dict[str, str]],
    sample_iris: dict[str, URIRef],
    as_of: date,
) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    missing_sheets = set(TRIP_SHEETS.values()) - set(workbook.sheetnames)
    if missing_sheets:
        raise ValueError(f"Missing pH workbook sheets: {sorted(missing_sheets)}")

    rows: list[dict[str, object]] = []
    for trip, sheet_name in TRIP_SHEETS.items():
        sheet = workbook[sheet_name]
        headers = [clean_text(sheet.cell(1, column).value) for column in range(1, 11)]
        if headers != EXPECTED_HEADERS:
            raise ValueError(f"Unexpected header in {sheet_name}: {headers}")
        for source_row in range(2, sheet.max_row + 1):
            values = [sheet.cell(source_row, column).value for column in range(1, 11)]
            if not any(value is not None for value in values):
                continue
            sample_id = clean_text(values[0])
            parsed_id = parse_sample_id(sample_id)
            workbook_site = int(values[1]) if is_number(values[1]) else None
            workbook_rep = int(values[3]) if is_number(values[3]) else None
            workbook_compartment = clean_text(values[2])
            normalized_date, date_status = parse_measurement_date(values[4], as_of)
            calibration = clean_text(values[5])
            slope = float(values[6]) if is_number(values[6]) else None
            ph10 = float(values[7]) if is_number(values[7]) else None
            ph = float(values[8]) if is_number(values[8]) else None
            ph_raw = clean_text(values[8])
            notes = clean_text(values[9])
            ledger = sample_ledger.get((trip, sample_id))
            specimen_iri = sample_iris.get(sample_id)

            reason_codes: list[str] = []
            metadata_match = True
            if parsed_id is None:
                metadata_match = False
                reason_codes.append("sample_id_unparseable")
            else:
                for field, observed, expected in (
                    ("trip", trip, parsed_id["trip"]),
                    ("site", workbook_site, parsed_id["site"]),
                    ("compartment", workbook_compartment, parsed_id["compartment"]),
                    ("replicate", workbook_rep, parsed_id["replicate"]),
                ):
                    if observed != expected:
                        metadata_match = False
                        reason_codes.append(f"{field}_conflicts_with_sample_id")
            if ledger is None:
                metadata_match = False
                reason_codes.append("absent_from_sample_ledger")
            else:
                if workbook_site is not None and clean_text(ledger["site"]) != str(workbook_site):
                    metadata_match = False
                    reason_codes.append("site_conflicts_with_sample_ledger")
                if workbook_compartment != ledger["compartment"]:
                    metadata_match = False
                    reason_codes.append("compartment_conflicts_with_sample_ledger")
            if specimen_iri is None:
                metadata_match = False
                reason_codes.append("absent_from_sample_abox")

            calibration_pass = calibration.lower() == "yes"
            slope_pass = slope is not None and 95.0 <= slope <= 102.0
            ph10_pass = ph10 is not None and abs(ph10 - 10.0) <= 0.1000001
            value_pass = ph is not None and 0.0 <= ph <= 14.0

            if ph is None:
                if notes.lower() == "depleted" or ph_raw in {"-", "–", "—"}:
                    disposition = "NOT_MEASURED_DEPLETED"
                    reason_codes.append("sample_depleted")
                elif ph10 is not None and 0.0 <= ph10 <= 14.0 and not ph10_pass:
                    disposition = "QUARANTINED_COLUMN_SHIFT_CANDIDATE"
                    reason_codes.append("numeric_value_in_ph10_column_without_ph_value")
                else:
                    disposition = "PENDING_MEASUREMENT"
                    reason_codes.append("ph_value_missing")
            elif not value_pass:
                disposition = "QUARANTINED_VALUE_RANGE"
                reason_codes.append("ph_outside_0_14")
            elif not metadata_match:
                disposition = "QUARANTINED_SPECIMEN_RECONCILIATION"
            elif not calibration_pass or not slope_pass or not ph10_pass:
                disposition = "QUARANTINED_QC"
                if not calibration_pass:
                    reason_codes.append("calibration_not_confirmed")
                if not slope_pass:
                    reason_codes.append("slope_outside_95_102")
                if not ph10_pass:
                    reason_codes.append("ph10_check_outside_9_9_10_1")
            elif date_status != "accepted":
                disposition = "QUARANTINED_DATE"
                reason_codes.append(f"measurement_date_{date_status}")
            else:
                disposition = ADMITTED

            accepted = disposition == ADMITTED
            rows.append(
                {
                    "source_sheet": sheet_name,
                    "source_row": source_row,
                    "trip": trip,
                    "sample_id": sample_id,
                    "specimen_iri": str(specimen_iri or ""),
                    "site": workbook_site,
                    "compartment": workbook_compartment,
                    "replicate": workbook_rep,
                    "measurement_date_raw": clean_text(values[4]),
                    "measurement_date": normalized_date,
                    "date_status": date_status,
                    "calibration_recorded": calibration,
                    "slope_percent": slope,
                    "ph10_check": ph10,
                    "ph_value_raw": ph_raw,
                    "ph_value": ph,
                    "notes": notes,
                    "metadata_match": metadata_match,
                    "calibration_pass": calibration_pass,
                    "slope_pass": slope_pass,
                    "ph10_check_pass": ph10_pass,
                    "value_range_pass": value_pass,
                    "disposition": disposition,
                    "ecology_eligible": accepted,
                    "kg_eligible": accepted,
                    "reason_codes": ";".join(dict.fromkeys(reason_codes)),
                }
            )
    return rows


def build_graph(
    rows: list[dict[str, object]],
    workbook_sha: str,
    dataset_version: str,
    dataset_status: str,
    dataset_purpose: str,
) -> tuple[Graph, list[dict[str, str]], list[dict[str, object]], URIRef]:
    graph = Graph()
    graph.bind("rak", BASE)
    graph.bind("sio", SIO)
    graph.bind("pato", PATO)
    graph.bind("uo", UO)
    graph.bind("dcterms", DCTERMS)

    version_slug = re.sub(r"[^a-z0-9]+", "_", dataset_version.lower()).strip("_")
    ontology_iri = URIRef(str(BASE) + f"rubalkhali_ph_{version_slug}.owl")
    graph.add((ontology_iri, RDF.type, OWL.Ontology))
    graph.add((ontology_iri, OWL.versionInfo, Literal(dataset_version)))
    graph.add(
        (
            ontology_iri,
            DCTERMS.description,
            Literal(
                f"Versioned pH module for {dataset_purpose}; status "
                f"{dataset_status}. Only audited rows are represented."
            ),
        )
    )

    dataset_iri, dataset_key = stable_iri("source", workbook_sha)
    graph.add((dataset_iri, RDF.type, DCAT.Dataset))
    graph.add((dataset_iri, DCTERMS.identifier, Literal(dataset_version)))
    graph.add((dataset_iri, DCTERMS.identifier, Literal(workbook_sha)))
    graph.add((dataset_iri, DCTERMS.title, Literal(f"Empty Quarter soil-pH workbook snapshot {dataset_version}")))
    graph.add((dataset_iri, DCTERMS.description, Literal(f"{dataset_status} source snapshot for {dataset_purpose}.")))

    protocol_iri, protocol_key = stable_iri("protocol", PROTOCOL_KEY)
    graph.add((protocol_iri, RDF.type, SIO.SIO_001043))
    graph.add((protocol_iri, DCTERMS.identifier, Literal(PROTOCOL_KEY)))
    graph.add((protocol_iri, RDFS.label, Literal("1:2.5 CaCl2 soil-pH protocol")))
    graph.add(
        (
            protocol_iri,
            DCTERMS.description,
            Literal(
                "Dry soil sieved below 2 mm and mixed without fine grinding; "
                "one independent 1 g scoop suspended in 2.5 mL 0.01 M CaCl2; "
                "two-point pH 7/10 calibration; pH 10 check accepted within "
                "0.1 of nominal; slope target 95-102%; ATC enabled."
            ),
        )
    )
    graph.add((protocol_iri, DCTERMS.source, dataset_iri))

    device_iri, device_key = stable_iri("device", "accumet-AB315-probe3")
    graph.add((device_iri, RDF.type, SIO.SIO_000956))
    graph.add((device_iri, DCTERMS.identifier, Literal("accumet-AB315-probe3")))
    graph.add((device_iri, RDFS.label, Literal("accumet AB315 with validated Probe 3")))

    registry: list[dict[str, str]] = []
    for iri, role, canonical in (
        (dataset_iri, "source", dataset_key),
        (protocol_iri, "protocol", protocol_key),
        (device_iri, "device", device_key),
    ):
        registry.append(
            {
                "entity_iri": str(iri),
                "entity_role": role,
                "canonical_key": canonical,
                "key_sha256": str(iri).rsplit("_", 1)[-1],
                "sample_id": "",
                "source_sheet": "",
                "source_row": "",
            }
        )

    accepted = [row for row in rows if row["disposition"] == ADMITTED]
    sessions: dict[tuple[object, ...], URIRef] = {}
    session_rows: list[dict[str, object]] = []
    for row in accepted:
        session_key_values = (
            row["trip"],
            row["measurement_date"],
            row["slope_percent"],
            row["ph10_check"],
        )
        if session_key_values not in sessions:
            session_key_text = "|".join(clean_text(value) for value in session_key_values)
            session_iri, session_key = stable_iri("session", session_key_text)
            sessions[session_key_values] = session_iri
            graph.add((session_iri, RDF.type, SIO.SIO_000994))
            graph.add((session_iri, DCTERMS.identifier, Literal(session_key_text)))
            graph.add((session_iri, RDFS.label, Literal(f"pH measurement session {session_key_text}")))
            graph.add((session_iri, DCTERMS.date, Literal(row["measurement_date"], datatype=XSD.date)))
            graph.add(
                (
                    session_iri,
                    DCTERMS.description,
                    Literal(
                        f"Workbook-reported electrode slope {row['slope_percent']}%; "
                        f"pH 10 check {row['ph10_check']}."
                    ),
                )
            )
            graph.add((session_iri, SIO.SIO_000132, device_iri))
            graph.add((session_iri, SIO.SIO_000772, dataset_iri))
            graph.add((dataset_iri, SIO.SIO_000773, session_iri))
            registry.append(
                {
                    "entity_iri": str(session_iri),
                    "entity_role": "session",
                    "canonical_key": session_key,
                    "key_sha256": str(session_iri).rsplit("_", 1)[-1],
                    "sample_id": "",
                    "source_sheet": row["source_sheet"],
                    "source_row": row["source_row"],
                }
            )
            session_rows.append(
                {
                    "session_iri": str(session_iri),
                    "trip": row["trip"],
                    "measurement_date": row["measurement_date"],
                    "slope_percent": row["slope_percent"],
                    "ph10_check": row["ph10_check"],
                    "protocol_iri": str(protocol_iri),
                    "device_iri": str(device_iri),
                }
            )

        base_key = "|".join(
            [
                f"trip{row['trip']}",
                str(row["sample_id"]),
                str(row["measurement_date"]),
                PROTOCOL_KEY,
            ]
        )
        process_iri, process_key = stable_iri("process", base_key)
        quality_iri, quality_key = stable_iri("quality", base_key)
        value_iri, value_key = stable_iri("value", base_key)
        specimen_iri = URIRef(str(row["specimen_iri"]))
        session_iri = sessions[session_key_values]
        label_tail = f"{row['sample_id']} on {row['measurement_date']}"

        graph.add((process_iri, RDF.type, SIO.SIO_001054))
        graph.add((process_iri, RDFS.label, Literal(f"Soil pH measuring process for {label_tail}")))
        graph.add((process_iri, DCTERMS.identifier, Literal(base_key)))
        graph.add((process_iri, DCTERMS.date, Literal(row["measurement_date"], datatype=XSD.date)))
        graph.add((process_iri, SIO.SIO_000230, specimen_iri))
        graph.add((process_iri, SIO.SIO_000291, specimen_iri))
        graph.add((process_iri, SIO.SIO_000229, value_iri))
        graph.add((process_iri, SIO.SIO_000339, protocol_iri))
        graph.add((process_iri, SIO.SIO_000132, device_iri))
        graph.add((process_iri, SIO.SIO_000068, session_iri))
        graph.add((process_iri, SIO.SIO_000772, dataset_iri))
        graph.add((dataset_iri, SIO.SIO_000773, process_iri))

        graph.add((quality_iri, RDF.type, PATO.PATO_0001842))
        graph.add((quality_iri, RDFS.label, Literal(f"Acidity quality of {row['sample_id']} under CaCl2 protocol")))
        graph.add((quality_iri, SIO.SIO_000011, specimen_iri))
        graph.add((quality_iri, SIO.SIO_000216, value_iri))
        graph.add((specimen_iri, SIO.SIO_000008, quality_iri))

        graph.add((value_iri, RDF.type, SIO.SIO_000070))
        graph.add((value_iri, RDF.type, SIO.SIO_001089))
        graph.add((value_iri, RDFS.label, Literal(f"pH value for {label_tail}")))
        graph.add((value_iri, DCTERMS.identifier, Literal(base_key)))
        graph.add((value_iri, SIO.SIO_000300, Literal(float(row["ph_value"]), datatype=XSD.double)))
        graph.add((value_iri, SIO.SIO_000221, UO.UO_0000196))
        graph.add((value_iri, SIO.SIO_000215, quality_iri))
        graph.add((value_iri, SIO.SIO_000232, process_iri))

        for iri, role, canonical in (
            (process_iri, "process", process_key),
            (quality_iri, "quality", quality_key),
            (value_iri, "value", value_key),
        ):
            registry.append(
                {
                    "entity_iri": str(iri),
                    "entity_role": role,
                    "canonical_key": canonical,
                    "key_sha256": str(iri).rsplit("_", 1)[-1],
                    "sample_id": row["sample_id"],
                    "source_sheet": row["source_sheet"],
                    "source_row": row["source_row"],
                }
            )

    return graph, registry, session_rows, ontology_iri


AUDIT_COLUMNS = [
    "source_sheet",
    "source_row",
    "trip",
    "sample_id",
    "specimen_iri",
    "site",
    "compartment",
    "replicate",
    "measurement_date_raw",
    "measurement_date",
    "date_status",
    "calibration_recorded",
    "slope_percent",
    "ph10_check",
    "ph_value_raw",
    "ph_value",
    "notes",
    "metadata_match",
    "calibration_pass",
    "slope_pass",
    "ph10_check_pass",
    "value_range_pass",
    "disposition",
    "ecology_eligible",
    "kg_eligible",
    "reason_codes",
]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument(
        "--project-root",
        type=Path,
        default=Path(__file__).resolve().parents[2],
    )
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--as-of", type=date.fromisoformat, required=True)
    parser.add_argument("--dataset-version", required=True)
    parser.add_argument(
        "--dataset-status",
        choices=("FROZEN", "DRAFT"),
        required=True,
    )
    parser.add_argument(
        "--dataset-purpose",
        choices=("ecology-manuscript", "data-paper", "shared-manuscripts"),
        required=True,
    )
    parser.add_argument("--source-complete", action="store_true")
    parser.add_argument("--measurement-campaign-closed", action="store_true")
    args = parser.parse_args()

    root = args.project_root.resolve()
    workbook_path = args.workbook.resolve()
    output_dir = args.output_dir.resolve()
    normalized_dir = output_dir / "normalized"
    kg_dir = output_dir / "kg"
    normalized_dir.mkdir(parents=True, exist_ok=True)
    kg_dir.mkdir(parents=True, exist_ok=True)

    workbook_sha = sha256_file(workbook_path)
    ledger = read_sample_ledger(root / "data/release/sample_ledger.tsv")
    sample_abox = root / "data/processed/semantics/ontology/rubalkhali_samples.owl"
    sample_iris = read_sample_iris(sample_abox)
    rows = audit_workbook(workbook_path, ledger, sample_iris, args.as_of)
    graph, registry, sessions, ontology_iri = build_graph(
        rows,
        workbook_sha,
        args.dataset_version,
        args.dataset_status,
        args.dataset_purpose,
    )

    audit_path = normalized_dir / "ph_observation_audit.tsv"
    accepted_path = normalized_dir / "ph_accepted_measurements.tsv"
    sessions_path = normalized_dir / "ph_measurement_sessions.tsv"
    registry_path = normalized_dir / "ph_entity_registry.tsv"
    graph_path = kg_dir / "rubalkhali_ph_measurements.ttl"
    graph_rdfxml_path = kg_dir / "rubalkhali_ph_measurements.owl"
    write_tsv(audit_path, AUDIT_COLUMNS, rows)
    write_tsv(
        accepted_path,
        AUDIT_COLUMNS,
        [row for row in rows if row["disposition"] == ADMITTED],
    )
    write_tsv(
        sessions_path,
        ["session_iri", "trip", "measurement_date", "slope_percent", "ph10_check", "protocol_iri", "device_iri"],
        sessions,
    )
    write_tsv(
        registry_path,
        ["entity_iri", "entity_role", "canonical_key", "key_sha256", "sample_id", "source_sheet", "source_row"],
        registry,
    )
    serialize_sorted_turtle(graph, graph_path)
    serialize_rdfxml(graph, graph_rdfxml_path)

    dispositions = Counter(str(row["disposition"]) for row in rows)
    per_trip: dict[str, dict[str, int]] = {}
    for trip in range(1, 6):
        trip_rows = [row for row in rows if row["trip"] == trip]
        per_trip[str(trip)] = {
            "declared_target_samples": DECLARED_SAMPLE_COUNTS[trip],
            "listed_rows": len(trip_rows),
            "numeric_ph_rows": sum(row["ph_value"] is not None for row in trip_rows),
            "admitted_rows": sum(row["disposition"] == ADMITTED for row in trip_rows),
            "pending_rows": sum(row["disposition"] == "PENDING_MEASUREMENT" for row in trip_rows),
            "quarantined_rows": sum(str(row["disposition"]).startswith("QUARANTINED") for row in trip_rows),
            "not_measured_depleted_rows": sum(row["disposition"] == "NOT_MEASURED_DEPLETED" for row in trip_rows),
        }
    summary = {
        "schema_version": SCHEMA_VERSION,
        "status": f"{args.dataset_status}_{args.dataset_purpose.upper().replace('-', '_')}",
        "dataset_version": args.dataset_version,
        "dataset_purpose": args.dataset_purpose,
        "source_complete": args.source_complete,
        "ontology_iri": str(ontology_iri),
        "as_of": args.as_of.isoformat(),
        "source": {
            "path": str(workbook_path.relative_to(root)) if workbook_path.is_relative_to(root) else str(workbook_path),
            "sha256": workbook_sha,
            "bytes": workbook_path.stat().st_size,
        },
        "protocol_key": PROTOCOL_KEY,
        "counts": {
            "source_rows": len(rows),
            "admitted_rows": dispositions[ADMITTED],
            "rdf_measurement_processes": len(list(graph.subjects(RDF.type, SIO.SIO_001054))),
            "rdf_ph_values": len(list(graph.subjects(RDF.type, SIO.SIO_001089))),
            "sessions": len(sessions),
            "dispositions": dict(sorted(dispositions.items())),
        },
        "per_trip": per_trip,
        "release_gate": {
            "canonical_kg_imported": False,
            "ecology_analysis_frozen": (
                args.dataset_status == "FROZEN"
                and args.dataset_purpose
                in {"ecology-manuscript", "shared-manuscripts"}
            ),
            "source_complete": args.source_complete,
            "future_versions_may_add_measurements": True,
        },
    }
    if args.dataset_purpose == "shared-manuscripts":
        summary["measurement_campaign_closed"] = args.measurement_campaign_closed
        summary["collection_status"] = (
            "COMPLETE_AS_AVAILABLE"
            if args.measurement_campaign_closed
            else "MEASUREMENT_CAMPAIGN_OPEN"
        )
        summary["release_gate"].update(
            {
                "canonical_kg_import_allowed": (
                    args.dataset_status == "FROZEN"
                    and args.measurement_campaign_closed
                ),
                "measurement_campaign_closed": args.measurement_campaign_closed,
            }
        )
    summary_path = output_dir / "summary.json"
    summary_path.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    version_manifest = workbook_path.parent / "manifest.json"
    version_registry = root / "data/metadata/samples/ph/version_registry.tsv"
    shape_path = root / "config/ph/ph_measurements.shex"
    input_paths = [
        workbook_path,
        root / "data/release/sample_ledger.tsv",
        sample_abox,
        version_manifest,
        version_registry,
        shape_path,
        Path(__file__).resolve(),
    ]
    manifest_paths = [
        *input_paths,
        audit_path,
        accepted_path,
        sessions_path,
        registry_path,
        graph_path,
        graph_rdfxml_path,
        summary_path,
    ]
    manifest_rows = []
    for path in manifest_paths:
        manifest_rows.append(
            {
                "path": str(path.relative_to(root)) if path.is_relative_to(root) else str(path),
                "sha256": sha256_file(path),
                "bytes": path.stat().st_size,
                "role": "input" if path in input_paths else "output",
            }
        )
    write_tsv(output_dir / "input_output_manifest.tsv", ["path", "sha256", "bytes", "role"], manifest_rows)
    print(json.dumps(summary, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
