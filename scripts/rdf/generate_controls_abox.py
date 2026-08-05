#!/usr/bin/env python3
"""Generate normalized control metadata and the SIO-aligned control ABox.

Human-readable EB labels are aliases only.  Every instance IRI is a SHA-256
digest of a typed immutable key.  Extraction batches can contain samples from
several trips and are never asserted as parts of expeditions.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import re
from collections import defaultdict
from pathlib import Path
from xml.sax.saxutils import escape, quoteattr

from openpyxl import load_workbook
from rdflib import BNode, DCTERMS, OWL, RDF, RDFS, XSD, Graph, Literal, Namespace, URIRef

BASE = Namespace("https://rubalkhali.science/kb/")
SIO = Namespace("http://semanticscience.org/resource/")
NCBITAXON = Namespace("http://purl.obolibrary.org/obo/NCBITaxon_")
UO = Namespace("http://purl.obolibrary.org/obo/UO_")
GROUND_TRUTH_SOURCE = (
    "data/metadata/samples/controls/control_ground_truth.tsv"
)
POSITIVE_RECOVERY_EVIDENCE = (
    "analysis/v3/control_audit/positive_control_expected_taxon_recovery.tsv"
)
DC = Namespace("http://purl.org/dc/elements/1.1/")
DCAT = Namespace("http://www.w3.org/ns/dcat#")
SCHEME_VERSION = "control-key-v1"

PRODUCT_TAXA = {
    "D6322": [
        ("287", "Pseudomonas aeruginosa", "14", "gDNA_percent"),
        ("562", "Escherichia coli", "14", "gDNA_percent"),
        ("28901", "Salmonella enterica", "14", "gDNA_percent"),
        ("1351", "Enterococcus faecalis", "14", "gDNA_percent"),
        ("1280", "Staphylococcus aureus", "14", "gDNA_percent"),
        ("1639", "Listeria monocytogenes", "14", "gDNA_percent"),
        ("1423", "Bacillus subtilis", "14", "gDNA_percent"),
        ("4932", "Saccharomyces cerevisiae", "2", "gDNA_percent"),
    ],
    "D6300": [
        ("287", "Pseudomonas aeruginosa", "12", "gDNA_percent"),
        ("562", "Escherichia coli", "12", "gDNA_percent"),
        ("28901", "Salmonella enterica", "12", "gDNA_percent"),
        ("1613", "Lactobacillus fermentum", "12", "gDNA_percent"),
        ("1351", "Enterococcus faecalis", "12", "gDNA_percent"),
        ("1280", "Staphylococcus aureus", "12", "gDNA_percent"),
        ("1639", "Listeria monocytogenes", "12", "gDNA_percent"),
        ("1423", "Bacillus subtilis", "12", "gDNA_percent"),
        ("4932", "Saccharomyces cerevisiae", "2", "gDNA_percent"),
        ("5207", "Cryptococcus neoformans", "2", "gDNA_percent"),
    ],
}

ROLE_CLASS = {
    "positive_microbiome_control": BASE.RAK_0000304,
    "negative_microbiome_control": BASE.RAK_0000305,
    "extraction_blank": BASE.RAK_0000306,
    "pcr_blank": BASE.RAK_0000307,
}
STATUS_CLASS = {
    "confirmed": BASE.RAK_0000314,
    "provisional": BASE.RAK_0000314,
    "unresolved": BASE.RAK_0000314,
    "rejected": BASE.RAK_0000314,
    "not_applicable": BASE.RAK_0000314,
}


def digest_id(kind: str, key_type: str, key: str) -> tuple[URIRef, str, str]:
    canonical = f"{kind}|{key_type}|{key}"
    digest = hashlib.sha256(canonical.encode("utf-8")).hexdigest()
    return URIRef(str(BASE) + "RAK_CTL_" + digest), canonical, digest


def split_ids(value: object) -> list[str]:
    return [part.strip() for part in str(value or "").split(",") if part.strip()]


def parse_index(path: str) -> str:
    match = re.search(r"(?:Index-Pair-|_(?:SU|UDP)0?)(\d+)", path)
    return match.group(1) if match else ""


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def write_tsv(path: Path, columns: list[str], rows: list[dict[str, object]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


RDFXML_NAMESPACES = (
    ("dc", "http://purl.org/dc/elements/1.1/"),
    ("dcterms", "http://purl.org/dc/terms/"),
    ("owl", "http://www.w3.org/2002/07/owl#"),
    ("rak", "https://rubalkhali.science/kb/"),
    ("rdf", "http://www.w3.org/1999/02/22-rdf-syntax-ns#"),
    ("rdfs", "http://www.w3.org/2000/01/rdf-schema#"),
    ("sio", "http://semanticscience.org/resource/"),
)


def rdfxml_qname(iri: URIRef) -> str:
    value = str(iri)
    for prefix, namespace in RDFXML_NAMESPACES:
        if value.startswith(namespace):
            return f"{prefix}:{value.removeprefix(namespace)}"
    raise ValueError(f"predicate has no deterministic RDF/XML prefix: {value}")


def serialize_deterministic_rdfxml(graph: Graph, path: Path) -> None:
    """Serialize this URI/literal-only ABox with stable triple ordering."""
    if any(isinstance(term, BNode) for triple in graph for term in triple):
        raise ValueError("deterministic control RDF/XML does not permit blank nodes")
    lines = ['<?xml version="1.0" encoding="utf-8"?>', "<rdf:RDF"]
    for prefix, namespace in RDFXML_NAMESPACES:
        lines.append(f"  xmlns:{prefix}={quoteattr(namespace)}")
    lines[-1] += ">"
    for subject in sorted(set(graph.subjects()), key=str):
        lines.append(f"  <rdf:Description rdf:about={quoteattr(str(subject))}>")
        triples = sorted(
            graph.predicate_objects(subject),
            key=lambda pair: (
                str(pair[0]),
                0 if isinstance(pair[1], URIRef) else 1,
                str(pair[1]),
                str(getattr(pair[1], "datatype", "") or ""),
                str(getattr(pair[1], "language", "") or ""),
            ),
        )
        for predicate, obj in triples:
            qname = rdfxml_qname(predicate)
            if isinstance(obj, URIRef):
                lines.append(
                    f"    <{qname} rdf:resource={quoteattr(str(obj))}/>"
                )
                continue
            attributes = ""
            if obj.language:
                attributes = f" xml:lang={quoteattr(obj.language)}"
            elif obj.datatype:
                attributes = f" rdf:datatype={quoteattr(str(obj.datatype))}"
            lines.append(
                f"    <{qname}{attributes}>{escape(str(obj))}</{qname}>"
            )
        lines.append("  </rdf:Description>")
    lines.append("</rdf:RDF>")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


class Builder:
    def __init__(self, root: Path) -> None:
        self.root = root
        self.graph = Graph()
        self.graph.bind("rak", BASE)
        self.graph.bind("sio", SIO)
        self.graph.bind("dcterms", DCTERMS)
        self.graph.bind("ncbitaxon", NCBITAXON)
        self.graph.bind("uo", UO)
        self.registry: list[dict[str, object]] = []
        self.aliases: list[dict[str, object]] = []
        self.roles: list[dict[str, object]] = []
        self.processes: list[dict[str, object]] = []
        self.occurrences: list[dict[str, object]] = []
        self.composition: list[dict[str, object]] = []
        self.assertions: list[dict[str, object]] = []
        self.dispositions: list[dict[str, object]] = []
        self.entities: dict[tuple[str, str, str], URIRef] = {}
        self.evidence: dict[str, URIRef] = {}
        self.statuses: dict[str, URIRef] = {}
        self.fields: dict[str, URIRef] = {}
        self.relations: dict[str, URIRef] = {}
        self.extraction_process_by_sample: dict[str, URIRef] = {}
        dna_graph = Graph()
        dna_graph.parse(root / "data/processed/semantics/ontology/rubalkhali_dna.owl")
        extraction_processes = set(dna_graph.subjects(RDF.type, BASE.RAK_0000309))
        extraction_processes.update(dna_graph.subjects(RDF.type, SIO.SIO_000994))
        for process in extraction_processes:
            for label in dna_graph.objects(process, RDFS.label):
                match = re.fullmatch(r"DNA extraction process for sample (.+)", str(label))
                if match:
                    self.extraction_process_by_sample[match.group(1)] = URIRef(process)

    def entity(
        self,
        kind: str,
        key_type: str,
        key: str,
        identifier: str,
        label: str,
        source: str,
        source_row: object = "",
        status: str = "confirmed",
    ) -> URIRef:
        lookup = (kind, key_type, key)
        if lookup in self.entities:
            return self.entities[lookup]
        iri, canonical, digest = digest_id(*lookup)
        self.entities[lookup] = iri
        self.registry.append(
            {
                "entity_id": str(iri),
                "entity_kind": kind,
                "stable_key_type": key_type,
                "stable_key": key,
                "iri_scheme_version": SCHEME_VERSION,
                "canonical_key": canonical,
                "key_sha256": digest,
                "primary_identifier": identifier,
                "primary_label": label,
                "source_snapshot_id": source,
                "source_row": source_row,
                "curation_status": status,
            }
        )
        self.graph.add((iri, DCTERMS.identifier, Literal(identifier)))
        self.graph.add((iri, RDFS.label, Literal(label)))
        return iri

    def add_alias(self, entity: URIRef, alias: str, alias_type: str, source: str, row: object) -> None:
        self.aliases.append(
            {
                "entity_id": str(entity),
                "alias": alias,
                "alias_type": alias_type,
                "source_snapshot_id": source,
                "source_row": row,
            }
        )

    def evidence_node(self, source: str) -> URIRef:
        if source in self.evidence:
            return self.evidence[source]
        node = self.entity(
            "evidence_record", "repository_path", source, source,
            f"Control evidence: {source}", source
        )
        self.graph.add((node, RDF.type, SIO.SIO_000089))
        self.evidence[source] = node
        return node

    def status_node(self, status: str) -> URIRef:
        if status in self.statuses:
            return self.statuses[status]
        node = self.entity(
            "metadata_status", "controlled_term", status, status,
            f"control metadata status: {status}", "CONTROL_KG_MODEL_DESIGN.md"
        )
        self.graph.add((node, RDF.type, STATUS_CLASS[status]))
        self.statuses[status] = node
        return node

    def field_node(self, field: str) -> URIRef:
        if field in self.fields:
            return self.fields[field]
        node = self.entity(
            "metadata_field", "controlled_term", field, field,
            f"control metadata field: {field}", "CONTROL_KG_MODEL_DESIGN.md"
        )
        self.graph.add((node, RDF.type, BASE.RAK_0000322))
        self.fields[field] = node
        return node

    def relation_node(self, predicate: URIRef) -> URIRef:
        value = str(predicate)
        if value in self.relations:
            return self.relations[value]
        node = self.entity(
            "relation_descriptor", "predicate_iri", value, value,
            f"asserted relation: {value}", "CONTROL_KG_MODEL_DESIGN.md"
        )
        self.graph.add((node, RDF.type, BASE.RAK_0000323))
        self.graph.add((node, BASE.RAK_2000098, Literal(value, datatype=XSD.anyURI)))
        self.relations[value] = node
        return node

    def assertion(
        self,
        subject: URIRef,
        predicate: URIRef,
        obj: URIRef,
        status: str,
        evidence: str,
        rationale: str,
        confidence: str = "high",
        promote: bool = True,
        assertion_class: URIRef | None = None,
    ) -> URIRef:
        key = f"{subject}|{predicate}|{obj}|{evidence}"
        node = self.entity(
            "metadata_assertion", "claim_tuple", key, hashlib.sha256(key.encode()).hexdigest(),
            "control metadata assertion", evidence
        )
        status_node = self.status_node(status)
        relation_node = self.relation_node(predicate)
        evidence_node = self.evidence_node(evidence)
        self.graph.add((node, RDF.type, BASE.RAK_0000313))
        if assertion_class is not None:
            self.graph.add((node, RDF.type, assertion_class))
        self.graph.add((node, BASE.RAK_2000093, subject))
        self.graph.add((node, BASE.RAK_2000094, relation_node))
        self.graph.add((node, BASE.RAK_2000095, obj))
        self.graph.add((node, BASE.RAK_2000090, status_node))
        self.graph.add((node, SIO.SIO_000772, evidence_node))
        self.graph.add((evidence_node, SIO.SIO_000773, node))
        self.graph.add((node, DCTERMS.description, Literal(rationale)))
        if status == "confirmed" and promote:
            self.graph.add((subject, predicate, obj))
            inverse = {
                SIO.SIO_000244: SIO.SIO_000245,
                SIO.SIO_000332: SIO.SIO_000629,
                SIO.SIO_000772: SIO.SIO_000773,
            }.get(predicate)
            if inverse is not None:
                self.graph.add((obj, inverse, subject))
        self.assertions.append(
            {
                "assertion_id": str(node),
                "subject_iri": str(subject),
                "predicate_iri": str(predicate),
                "object_iri": str(obj),
                "object_literal": "",
                "object_datatype": "",
                "status": status,
                "confidence": confidence,
                "evidence_id": str(evidence_node),
                "rationale": rationale,
            }
        )
        return node

    def disposition(
        self, about: URIRef, field: str, disposition: str, evidence: str, description: str
    ) -> None:
        key = f"{about}|{field}|{disposition}|{evidence}"
        node = self.entity(
            "metadata_disposition", "disposition_tuple", key,
            hashlib.sha256(key.encode()).hexdigest(), f"{disposition}: {field}", evidence
        )
        cls = BASE.RAK_0000320 if disposition == "unresolved" else BASE.RAK_0000321
        self.graph.add((node, RDF.type, cls))
        self.graph.add((node, SIO.SIO_000332, about))
        self.graph.add((about, SIO.SIO_000629, node))
        self.graph.add((node, BASE.RAK_2000097, self.field_node(field)))
        self.graph.add((node, BASE.RAK_2000090, self.status_node(
            "unresolved" if disposition == "unresolved" else "not_applicable"
        )))
        evidence_node = self.evidence_node(evidence)
        self.graph.add((node, SIO.SIO_000772, evidence_node))
        self.graph.add((evidence_node, SIO.SIO_000773, node))
        self.graph.add((node, DCTERMS.description, Literal(description)))
        self.dispositions.append(
            {
                "disposition_id": str(node),
                "about_entity_id": str(about),
                "metadata_field": field,
                "disposition": disposition,
                "evidence_id": str(evidence_node),
                "description": description,
            }
        )

    def add_role(
        self, material: URIRef, role_type: str, process: URIRef | None,
        source: str, key: str, status: str = "confirmed"
    ) -> URIRef:
        role = self.entity(
            "control_role", "material_and_role", f"{material}|{role_type}",
            f"{key}:{role_type}", f"{role_type.replace('_', ' ')} for {key}", source
        )
        self.graph.add((role, RDF.type, ROLE_CLASS[role_type]))
        self.graph.add((material, SIO.SIO_000228, role))
        self.graph.add((role, SIO.SIO_000227, material))
        if process is not None:
            self.graph.add((process, SIO.SIO_000355, role))
            self.graph.add((role, SIO.SIO_000356, process))
        self.roles.append(
            {
                "role_id": str(role),
                "bearer_material_id": str(material),
                "role_type": role_type,
                "realized_in_process_id": str(process or ""),
                "assertion_status": status,
                "evidence_id": str(self.evidence_node(source)),
            }
        )
        return role

    def extraction_lineage(
        self, material: URIRef, key: str, source: str, row: object,
        batch: URIRef | None = None, process_date: str = ""
    ) -> tuple[URIRef, URIRef]:
        process = self.entity(
            "laboratory_process", "control_extraction", key,
            f"{key}:extraction", f"DNA extraction process for {key}", source, row
        )
        extract = self.entity(
            "dna_extract", "control_extraction_output", key,
            f"{key}:DNA", f"DNA extract of {key}", source, row
        )
        self.graph.add((process, RDF.type, BASE.RAK_0000309))
        self.graph.add((process, RDF.type, SIO.SIO_000994))
        self.graph.add((extract, RDF.type, BASE.RAK_0000040))
        for predicate, subject, obj in [
            (SIO.SIO_000230, process, material), (SIO.SIO_000231, material, process),
            (SIO.SIO_000291, process, material), (SIO.SIO_000292, material, process),
            (SIO.SIO_000229, process, extract), (SIO.SIO_000232, extract, process),
        ]:
            self.graph.add((subject, predicate, obj))
        if batch:
            self.graph.add((process, SIO.SIO_000068, batch))
            self.graph.add((batch, SIO.SIO_000028, process))
        self.processes.append(
            {
                "process_id": str(process),
                "process_type": "dna_extraction",
                "batch_id": str(batch or ""),
                "input_entity_id": str(material),
                "output_entity_id": str(extract),
                "process_date": process_date,
                "protocol_id": "",
                "kit_id": "",
                "agent_id": "",
                "source_snapshot_id": source,
                "source_row": row,
            }
        )
        return process, extract

    def sequence_lineage(
        self, material_or_extract: URIRef, sample_id: str, source: str, source_row: object,
        read1: str, read2: str, sra: dict[str, str] | None = None
    ) -> None:
        library = self.entity(
            "amplicon_library", "facility_sample_id", sample_id,
            sra.get("library_name", sample_id + "_amplicon_library") if sra else sample_id,
            f"16S amplicon library for {sample_id}", source, source_row
        )
        pcr = self.entity(
            "laboratory_process", "amplicon_pcr", sample_id,
            sample_id + ":PCR", f"amplicon PCR process for {sample_id}", source, source_row
        )
        run_key = (sra or {}).get("run_accession") or ("IBEX:" + sample_id)
        run = self.entity(
            "sequencing_run", "run_accession_or_facility_sample", run_key,
            run_key, f"control sequencing assay for {sample_id}", source, source_row
        )
        fastq = self.entity(
            "fastq_dataset", "run_accession_or_facility_sample", run_key,
            run_key, f"control-derived FASTQ dataset for {sample_id}", source, source_row
        )
        self.graph.add((library, RDF.type, BASE.RAK_0000060))
        self.graph.add((pcr, RDF.type, BASE.RAK_0000311))
        self.graph.add((pcr, RDF.type, SIO.SIO_000994))
        self.graph.add((run, RDF.type, BASE.RAK_0000312))
        self.graph.add((fastq, RDF.type, BASE.RAK_0000063))
        self.graph.add((fastq, RDF.type, BASE.RAK_0000318))
        for predicate, subject, obj in [
            (SIO.SIO_000230, pcr, material_or_extract),
            (SIO.SIO_000231, material_or_extract, pcr),
            (SIO.SIO_000229, pcr, library),
            (SIO.SIO_000232, library, pcr),
            (SIO.SIO_000230, run, library),
            (SIO.SIO_000231, library, run),
            (SIO.SIO_000229, run, fastq),
            (SIO.SIO_000232, fastq, run),
        ]:
            self.graph.add((subject, predicate, obj))
        if sra:
            self.graph.add((library, RDFS.seeAlso, URIRef(
                "https://identifiers.org/biosample/" + sra["biosample_accession"]
            )))
            self.graph.add((fastq, RDFS.seeAlso, URIRef(
                "https://identifiers.org/insdc.run/" + sra["run_accession"]
            )))
        self.graph.add((fastq, DC.source, Literal(read1)))
        self.graph.add((fastq, DC.source, Literal(read2)))
        self.processes.extend([
            {
                "process_id": str(pcr), "process_type": "amplicon_pcr",
                "batch_id": "", "input_entity_id": str(material_or_extract),
                "output_entity_id": str(library), "process_date": "",
                "protocol_id": "", "kit_id": "", "agent_id": "",
                "source_snapshot_id": source, "source_row": source_row,
            },
            {
                "process_id": str(run), "process_type": "control_sequencing_assay",
                "batch_id": "", "input_entity_id": str(library),
                "output_entity_id": str(fastq), "process_date": "",
                "protocol_id": "", "kit_id": "", "agent_id": "",
                "source_snapshot_id": source, "source_row": source_row,
            },
        ])
        self.occurrences.append(
            {
                "library_entity_id": str(library),
                "run_entity_id": str(run),
                "fastq_entity_id": str(fastq),
                "library_accession": (sra or {}).get("experiment_accession", ""),
                "biosample_accession": (sra or {}).get("biosample_accession", ""),
                "run_accession": (sra or {}).get("run_accession", ""),
                "facility_run_id": run_key if not sra else "",
                "index_code": parse_index(read1),
                "index_sequence": "",
                "lane": (re.search(r"_L(\d{3})_", read1) or ["", ""])[1],
                "read_1_file": read1,
                "read_2_file": read2,
                "source_snapshot_id": source,
                "source_row": source_row,
            }
        )

    def add_products(self) -> dict[str, URIRef]:
        products = {}
        sources = {
            "D6322": "https://www.zymoresearch.com/products/zymobiomics-hmw-dna-standard",
            "D6300": "https://www.zymoresearch.com/products/zymobiomics-microbial-community-standard",
        }
        for product, taxa in PRODUCT_TAXA.items():
            source = sources[product]
            product_node = self.entity(
                "control_material", "catalogue_number", product, product,
                f"ZymoBIOMICS {'HMW DNA' if product == 'D6322' else 'Microbial Community'} Standard ({product})",
                source
            )
            spec = self.entity(
                "composition_specification", "catalogue_number_and_version",
                product + "|manufacturer_snapshot_2026-07-30", product + ":composition",
                f"manufacturer composition specification for {product}", source
            )
            self.graph.add((product_node, RDF.type, BASE.RAK_0000301))
            self.graph.add((spec, RDF.type, BASE.RAK_0000316))
            self.graph.add((product_node, SIO.SIO_000339, spec))
            self.graph.add((spec, SIO.SIO_000338, product_node))
            self.graph.add((spec, DCTERMS.source, URIRef(source)))
            for taxon_id, label, value, basis in taxa:
                taxon = NCBITAXON[taxon_id]
                expected_assertion = self.assertion(
                    spec,
                    BASE.RAK_2000092,
                    taxon,
                    "confirmed",
                    source,
                    (
                        f"Manufacturer specification for {product} lists "
                        f"{label} as an expected organism."
                    ),
                    assertion_class=BASE.RAK_0000317,
                )
                self.composition.append(
                    {
                        "composition_specification_id": str(spec),
                        "control_material_or_product_id": str(product_node),
                        "taxon_iri": str(taxon),
                        "expected_value": value,
                        "expected_unit_iri": str(UO["0000187"]),
                        "composition_basis": basis,
                        "product": product,
                        "lot": "",
                        "version": "manufacturer_snapshot_2026-07-30",
                        "assertion_status": "confirmed",
                        "assertion_id": str(expected_assertion),
                        "evidence_id": str(self.evidence_node(source)),
                    }
                )
            products[product] = product_node
        return products

    def build_trip5(self) -> None:
        source = "data/metadata/samples/Sequenced_Samples_by_EB_FifthTrip.xlsx"
        sample_source = "data/metadata/samples/controls/source_snapshots/ibex_trip5_16s_samplesheet.tsv"
        ws = load_workbook(self.root / source, read_only=True, data_only=True).active
        mapped: dict[str, dict[str, object]] = {}
        for row_no, row in enumerate(ws.iter_rows(values_only=True), start=1):
            label = str(row[0] or "").strip()
            if label.startswith("EB") and label[2:].isdigit() and 1 <= int(label[2:]) <= 17:
                mapped[label] = {
                    "date": __import__("datetime").datetime.strptime(str(row[1]), "%m/%d/%Y").date().isoformat(),
                    "index": str(row[2]).strip().replace("  ", " "),
                    "samples": split_ids(row[4]),
                    "row": row_no,
                }
        sra_rows = {
            row["sample_name"]: row for row in read_tsv(
                self.root / "data/metadata/sra-submissions/submission-sheet.tsv"
            )
        }
        sheet_rows = {
            row["sampleID"]: row for row in read_tsv(self.root / sample_source)
            if row["sampleID"].startswith("EB") or row["sampleID"].startswith("Negative")
        }
        for sample_id, row in sheet_rows.items():
            key = f"trip5_16S|{sample_id}|{parse_index(row['forwardReads'])}"
            material = self.entity(
                "control_material", "assay_sample_and_index", key, sample_id,
                f"Trip 5 16S extraction blank {sample_id}", sample_source, sample_id
            )
            self.add_alias(material, sample_id, "laboratory_label", sample_source, sample_id)
            self.graph.add((material, RDF.type, BASE.RAK_0000300))
            self.graph.add((material, RDF.type, BASE.RAK_0000302))
            batch = None
            date = ""
            if sample_id in mapped:
                info = mapped[sample_id]
                date = str(info["date"])
                batch = self.entity(
                    "batch", "extraction_date_and_blank_index",
                    f"{date}|{info['index']}", f"extraction:{date}",
                    f"extraction batch {date} ({sample_id})", source, info["row"]
                )
                self.graph.add((batch, RDF.type, BASE.RAK_0000308))
            process, extract = self.extraction_lineage(
                material, key, source if sample_id in mapped else sample_source,
                mapped.get(sample_id, {}).get("row", sample_id), batch, date
            )
            if batch:
                for biological_sample in mapped[sample_id]["samples"]:
                    biological_process = self.extraction_process_by_sample.get(biological_sample)
                    if biological_process:
                        self.graph.add((biological_process, SIO.SIO_000068, batch))
                        self.graph.add((batch, SIO.SIO_000028, biological_process))
                        self.processes.append(
                            {
                                "process_id": str(biological_process),
                                "process_type": "biological_dna_extraction_membership",
                                "batch_id": str(batch),
                                "input_entity_id": f"sample_identifier:{biological_sample}",
                                "output_entity_id": "",
                                "process_date": date,
                                "protocol_id": "", "kit_id": "", "agent_id": "",
                                "source_snapshot_id": source,
                                "source_row": mapped[sample_id]["row"],
                            }
                        )
            self.add_role(material, "extraction_blank", process, sample_source, sample_id)
            self.sequence_lineage(
                extract, sample_id, sample_source, sample_id,
                row["forwardReads"], row["reverseReads"], sra_rows.get(sample_id)
            )
            if sample_id not in mapped:
                self.disposition(
                    material, "extraction_batch_membership", "unresolved", sample_source,
                    "The library is an author-confirmed extraction blank, but no extraction-day/sample map is frozen."
                )

    def build_trip4(self) -> None:
        source = "data/metadata/samples/EB_Sample_Map_FourthTrip2.xlsx"
        fastq_source = "data/metadata/samplesheets/additional_fastqs_v2.tsv"
        ws = load_workbook(self.root / source, read_only=True, data_only=True).active
        workbook_rows = list(ws.iter_rows(values_only=True))
        membership_count: defaultdict[str, int] = defaultdict(int)
        for row in workbook_rows:
            label = str(row[0] or "").strip()
            if label == "EB" or (label.startswith("EB") and label[2:].isdigit()):
                for sample in split_ids(row[4]):
                    membership_count[sample] += 1
        manifest_names = {
            "EB": "SEB",
            **{f"EB{i}": f"EB{i}" for i in range(1, 6)},
        }
        fastq_rows: dict[str, tuple[str, str, int]] = {}
        for manifest_row, record in enumerate(
            read_tsv(self.root / fastq_source), start=2
        ):
            sample_name = record["Sample Name"]
            if sample_name not in set(manifest_names.values()):
                continue
            paths = [part.strip() for part in record["FASTQ Paths"].split(";")]
            if (
                len(paths) != 2
                or "_R1_" not in paths[0]
                or "_R2_" not in paths[1]
            ):
                raise ValueError(
                    f"{fastq_source}:{manifest_row}: invalid paired FASTQ paths"
                )
            fastq_rows[sample_name] = (paths[0], paths[1], manifest_row)
        if set(fastq_rows) != set(manifest_names.values()):
            raise ValueError(
                "Trip-4 control FASTQs are incomplete in additional_fastqs_v2.tsv"
            )
        duplicate_dispositions: set[str] = set()
        for row_no, row in enumerate(workbook_rows, start=1):
            label = str(row[0] or "").strip()
            if label not in manifest_names:
                continue
            index_text = str(row[1]).strip()
            key = f"trip4_16S|{label}|{index_text}"
            material = self.entity(
                "control_material", "assay_sample_and_index", key, label,
                f"Trip 4 extraction blank {label}", source, row_no
            )
            self.add_alias(material, label, "laboratory_label", source, row_no)
            self.graph.add((material, RDF.type, BASE.RAK_0000300))
            self.graph.add((material, RDF.type, BASE.RAK_0000302))
            batch = self.entity(
                "batch", "workbook_row_and_index", f"{source}|{row_no}|{index_text}",
                f"trip4-extraction:{index_text}", f"Trip 4 extraction batch represented by {label}",
                source, row_no
            )
            self.graph.add((batch, RDF.type, BASE.RAK_0000308))
            process, extract = self.extraction_lineage(material, key, source, row_no, batch)
            self.add_role(material, "extraction_blank", process, source, label)
            read1, read2, manifest_row = fastq_rows[manifest_names[label]]
            self.sequence_lineage(
                extract,
                "Trip4_" + label,
                fastq_source,
                manifest_row,
                read1,
                read2,
            )
            for sample in split_ids(row[4]):
                biological_process = self.extraction_process_by_sample.get(sample)
                if biological_process:
                    if membership_count[sample] > 1:
                        self.assertion(
                            biological_process, SIO.SIO_000068, batch, "provisional",
                            source,
                            f"Workbook row {row_no} assigns {sample} to this batch, but the same sample occurs in another EB row.",
                            confidence="low", promote=False,
                        )
                        if sample not in duplicate_dispositions:
                            self.disposition(
                                biological_process, "extraction_batch_membership",
                                "unresolved", source,
                                f"{sample} is mapped to more than one Trip-4 extraction blank; no batch link is promoted."
                            )
                            duplicate_dispositions.add(sample)
                    else:
                        self.graph.add((biological_process, SIO.SIO_000068, batch))
                        self.graph.add((batch, SIO.SIO_000028, biological_process))
                        self.processes.append(
                            {
                                "process_id": str(biological_process),
                                "process_type": "biological_dna_extraction_membership",
                                "batch_id": str(batch),
                                "input_entity_id": f"sample_identifier:{sample}",
                                "output_entity_id": "",
                                "process_date": "",
                                "protocol_id": "", "kit_id": "", "agent_id": "",
                                "source_snapshot_id": source, "source_row": row_no,
                            }
                        )
        # One physical extraction blank existed for the three-day extraction
        # but was not sequenced.
        affected = "22,28,35,43,53,46,13,25,54,55,44,38,41,33,2,27,51,16,59,57,56,49,45"
        material = self.entity(
            "control_material", "workbook_unsequenced_row", source + "|9",
            "Trip4-unsequenced-EB", "unsequenced Trip 4 three-day extraction blank", source, 9
        )
        self.add_alias(material, "EB", "laboratory_label", source, 9)
        self.graph.add((material, RDF.type, BASE.RAK_0000300))
        self.graph.add((material, RDF.type, BASE.RAK_0000302))
        batch = self.entity(
            "batch", "workbook_unsequenced_row", source + "|9",
            "trip4-three-day-extraction", "Trip 4 three-day extraction batch without a sequenced blank",
            source, 9
        )
        self.graph.add((batch, RDF.type, BASE.RAK_0000308))
        process, _ = self.extraction_lineage(material, source + "|9", source, 9, batch)
        self.add_role(material, "extraction_blank", process, source, "Trip4-unsequenced-EB")
        self.disposition(
            material, "control_sequence_occurrence", "unresolved", source,
            f"The blank was not sequenced; affected site identifiers: {affected}."
        )
        # Trip 4 had no positive control: absence is a disposition, not a fake
        # positive-control material.
        inventory = self.evidence_node(GROUND_TRUTH_SOURCE)
        self.disposition(
            inventory, "trip4_positive_control", "not_applicable",
            GROUND_TRUTH_SOURCE,
            "Author/laboratory record confirms that no positive control travelled with Trip 4."
        )
        # The workbook records a PCR blank at index 275 but no sequenced count
        # or FASTQ.  Preserve the physical/process record without fabricating a
        # sequence occurrence.
        pcr_material = self.entity(
            "control_material", "workbook_row_and_index", source + "|8|275",
            "Trip4-PCR-blank", "Trip 4 PCR blank", source, 8
        )
        self.add_alias(pcr_material, "PCR Blank", "laboratory_label", source, 8)
        pcr_process = self.entity(
            "laboratory_process", "pcr_blank_reaction", source + "|8|275",
            "Trip4-PCR-blank:PCR", "amplicon PCR process for Trip 4 PCR blank", source, 8
        )
        pcr_library = self.entity(
            "amplicon_library", "workbook_row_and_index", source + "|8|275",
            "Trip4-PCR-blank-library", "Trip 4 PCR blank reaction product", source, 8
        )
        self.graph.add((pcr_material, RDF.type, BASE.RAK_0000300))
        self.graph.add((pcr_material, RDF.type, BASE.RAK_0000302))
        self.graph.add((pcr_process, RDF.type, BASE.RAK_0000311))
        self.graph.add((pcr_process, RDF.type, SIO.SIO_000994))
        self.graph.add((pcr_library, RDF.type, BASE.RAK_0000060))
        self.graph.add((pcr_process, SIO.SIO_000230, pcr_material))
        self.graph.add((pcr_material, SIO.SIO_000231, pcr_process))
        self.graph.add((pcr_process, SIO.SIO_000229, pcr_library))
        self.graph.add((pcr_library, SIO.SIO_000232, pcr_process))
        self.add_role(pcr_material, "pcr_blank", pcr_process, source, "Trip4-PCR-blank")
        self.processes.append(
            {
                "process_id": str(pcr_process), "process_type": "amplicon_pcr",
                "batch_id": "", "input_entity_id": str(pcr_material),
                "output_entity_id": str(pcr_library), "process_date": "",
                "protocol_id": "", "kit_id": "", "agent_id": "",
                "source_snapshot_id": source, "source_row": 8,
            }
        )
        self.disposition(
            pcr_material, "control_sequence_occurrence", "unresolved", source,
            "The workbook records the PCR blank and index 275 but no sequenced count or FASTQ."
        )

    def build_july_controls(self, products: dict[str, URIRef]) -> None:
        source = "data/metadata/samples/controls/source_snapshots/ibex_20250714_qiime2/controls-metadata.tsv"
        rows = read_tsv(self.root / source)
        # FPosCtrl3 was present in the frozen samplesheet/table but omitted
        # from the legacy metadata; add it explicitly from that source.
        july_sheet = "data/metadata/samples/controls/source_snapshots/ibex_20250714_16s_samplesheet.tsv"
        sheet_rows = {row["sampleID"]: row for row in read_tsv(self.root / july_sheet)}
        known = {row["sampleID"] for row in rows}
        if "e8665_FPosCtrl3" in sheet_rows and "e8665_FPosCtrl3" not in known:
            r = sheet_rows["e8665_FPosCtrl3"]
            rows.append({
                "sampleID": "e8665_FPosCtrl3", "trip": "3", "replicate": "3",
                "compartment": "pos", "type": "control", "control": "yes",
                "file_R1": r["forwardReads"], "file_R2": r["reverseReads"],
            })
        positive_products = {
            "e0553_Ctrl_2": "D6322", "e0554_Ctrl_3": "D6322",
            "e0872_TMC1": "D6322",
            "e0874_Positive": "D6322", "e8294_FPosCtrl1": "D6300",
            "e8661_FPosCtrl2": "D6300", "e8665_FPosCtrl3": "D6300",
        }
        for row_no, row in enumerate(rows, start=2):
            sample_id = row["sampleID"]
            key = f"20250714_16S|{sample_id}|{parse_index(row['file_R1'])}"
            positive = sample_id in positive_products
            material = self.entity(
                "control_material", "assay_sample_and_index", key, sample_id,
                f"16S {'positive control' if positive else 'negative control'} {sample_id}",
                source, row_no
            )
            self.add_alias(material, sample_id, "laboratory_label", source, row_no)
            self.graph.add((material, RDF.type, BASE.RAK_0000300))
            self.graph.add((material, RDF.type, BASE.RAK_0000301 if positive else BASE.RAK_0000302))
            stage_input = material
            process = None
            if positive:
                product = positive_products[sample_id]
                self.assertion(
                    material, SIO.SIO_000244, products[product], "provisional",
                    GROUND_TRUTH_SOURCE,
                    (
                        f"The library is positive-labelled and the trip-level "
                        f"design used {product}, but the library-level product "
                        "mapping is not author-confirmed."
                    ),
                    confidence="low",
                    promote=False,
                )
                self.disposition(
                    material,
                    "positive_control_product_identity",
                    "unresolved",
                    POSITIVE_RECOVERY_EVIDENCE,
                    (
                        f"Expected-taxon recovery is evaluated against {product} "
                        "as a trip-design hypothesis; it does not establish that "
                        f"{sample_id} was derived from that product."
                    ),
                )
                if product == "D6300":
                    process, stage_input = self.extraction_lineage(material, key, source, row_no)
                role_type = "positive_microbiome_control"
            elif "PCR" in sample_id or "NTC" in sample_id:
                role_type = "pcr_blank"
            elif "Extraction" in sample_id:
                role_type = "extraction_blank"
                process, stage_input = self.extraction_lineage(material, key, source, row_no)
            else:
                role_type = "negative_microbiome_control"
                self.disposition(
                    material, "negative_control_processing_stage", "unresolved", source,
                    "The frozen label establishes a negative control but does not distinguish extraction blank from PCR blank."
                )
                if sample_id == "e0323_Ctrl_1_Trip1":
                    self.disposition(
                        material,
                        "control_polarity_reconciliation",
                        "unresolved",
                        source,
                        "The frozen control metadata labels e0323_Ctrl_1_Trip1 as negative. Its ambiguous numbered label is not sufficient to map it to the author-confirmed Trip-1 D6322 positive-control replicate set.",
                    )
            self.sequence_lineage(stage_input, sample_id, source, row_no, row["file_R1"], row["file_R2"])
            if role_type == "pcr_blank":
                # Reuse the PCR process created by sequence_lineage.
                pcr = self.entities[("laboratory_process", "amplicon_pcr", sample_id)]
                process = pcr
            self.add_role(material, role_type, process, source, sample_id)

    def build_trip5_shotgun(self, products: dict[str, URIRef]) -> None:
        source = GROUND_TRUTH_SOURCE
        run_id = "20260402_LL00134_0018_A23J2Y7LT4"
        assay = self.entity(
            "sequencing_run", "facility_run_id", run_id, run_id,
            "Trip 5 shotgun control sequencing assay", source
        )
        self.graph.add((assay, RDF.type, BASE.RAK_0000312))
        for role_name, identifier in [
            ("positive_microbiome_control", "Trip5-shotgun-positive"),
            ("extraction_blank", "Trip5-shotgun-negative"),
        ]:
            material = self.entity(
                "control_material", "facility_run_and_role", f"{run_id}|{role_name}",
                identifier, f"{identifier} control material", source
            )
            self.add_alias(material, identifier, "curated_run_role_label", source, "")
            self.graph.add((material, RDF.type, BASE.RAK_0000300))
            self.graph.add((
                material, RDF.type,
                BASE.RAK_0000301 if role_name == "positive_microbiome_control" else BASE.RAK_0000302
            ))
            extraction, extract = self.extraction_lineage(
                material, f"{run_id}|{role_name}", source, ""
            )
            self.add_role(material, role_name, extraction, source, identifier)
            self.graph.add((assay, SIO.SIO_000230, extract))
            self.graph.add((extract, SIO.SIO_000231, assay))
            if role_name == "positive_microbiome_control":
                self.assertion(
                    material, SIO.SIO_000244, products["D6300"], "confirmed", source,
                    "Author/laboratory record assigns the D6300 whole-cell standard to Trip 5."
                )
            self.disposition(
                material, "control_sequence_occurrence", "unresolved", source,
                f"The facility run is known ({run_id}), but the exact library identifier, FASTQ path and accession are not frozen."
            )
        self.disposition(
            assay, "amplicon_assay_applicability", "not_applicable", source,
            "This control pair was sequenced by shotgun metagenomics and is not a Trip-5 16S positive-control observation."
        )

    def add_sterile_bag_disposition(self) -> None:
        inventory = self.evidence_node(GROUND_TRUTH_SOURCE)
        self.disposition(
            inventory, "sterile_bag_field_control_inventory", "unresolved",
            GROUND_TRUTH_SOURCE,
            "A sterile-bag field-control inventory was requested, but no bag identifiers, deployment processes, or sequence occurrences were supplied; no instances were invented."
        )

    def save(self) -> None:
        normalized = self.root / "data/processed/metadata/controls"
        normalized.mkdir(parents=True, exist_ok=True)
        write_tsv(normalized / "control_entity_registry.tsv", list(self.registry[0]), self.registry)
        write_tsv(normalized / "control_aliases.tsv",
                  ["entity_id", "alias", "alias_type", "source_snapshot_id", "source_row"], self.aliases)
        write_tsv(normalized / "control_roles.tsv", list(self.roles[0]), self.roles)
        write_tsv(normalized / "laboratory_processes.tsv", list(self.processes[0]), self.processes)
        write_tsv(normalized / "control_sequence_occurrences.tsv", list(self.occurrences[0]), self.occurrences)
        write_tsv(normalized / "control_composition.tsv", list(self.composition[0]), self.composition)
        write_tsv(normalized / "control_assertions.tsv", list(self.assertions[0]), self.assertions)
        write_tsv(normalized / "control_metadata_dispositions.tsv", list(self.dispositions[0]), self.dispositions)
        ontology = URIRef(str(BASE) + "rubalkhali_controls.owl")
        self.graph.add((ontology, RDF.type, OWL.Ontology))
        self.graph.add((ontology, OWL.imports, URIRef(str(BASE))))
        self.graph.add((ontology, DCTERMS.title, Literal("Rub al-Khali laboratory control ABox")))
        out = self.root / "data/processed/semantics/ontology"
        self.graph.serialize(out / "rubalkhali_controls.ttl", format="turtle")
        serialize_deterministic_rdfxml(
            self.graph, out / "rubalkhali_controls.owl"
        )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--project-root",
        type=Path,
        default=Path(__file__).resolve().parents[2],
        help="Project tree containing data inputs and receiving generated outputs.",
    )
    args = parser.parse_args()
    root = args.project_root.resolve()
    builder = Builder(root)
    products = builder.add_products()
    builder.build_july_controls(products)
    builder.build_trip4()
    builder.build_trip5()
    builder.build_trip5_shotgun(products)
    builder.add_sterile_bag_disposition()
    builder.save()
    print(
        f"Generated {len(builder.graph):,} triples, {len(builder.registry)} registered "
        f"entities, {len(builder.roles)} roles and {len(builder.occurrences)} sequence occurrences."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
