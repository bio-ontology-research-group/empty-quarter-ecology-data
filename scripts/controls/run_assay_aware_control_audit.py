#!/usr/bin/env python3
"""Assay- and process-aware control audit for the Empty Quarter 16S data.

This intentionally does not reproduce the retired pooled filter.  Positive
controls are evaluation-only.  The primary contaminant screen uses only the
17 Trip-5 extraction blanks with frozen extraction-day/sample mappings and is
applied only to the compatible Trip-5 biological profiles.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import io
import json
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
from biom import load_table
from openpyxl import load_workbook
from scipy.stats import fisher_exact


EXPECTED = {
    "D6322": [
        ("Pseudomonas aeruginosa", "Pseudomonas", "bacterium", 14.0, "gDNA_percent"),
        ("Escherichia coli", "Escherichia", "bacterium", 14.0, "gDNA_percent"),
        ("Salmonella enterica", "Salmonella", "bacterium", 14.0, "gDNA_percent"),
        ("Enterococcus faecalis", "Enterococcus", "bacterium", 14.0, "gDNA_percent"),
        ("Staphylococcus aureus", "Staphylococcus", "bacterium", 14.0, "gDNA_percent"),
        ("Listeria monocytogenes", "Listeria", "bacterium", 14.0, "gDNA_percent"),
        ("Bacillus subtilis", "Bacillus", "bacterium", 14.0, "gDNA_percent"),
        ("Saccharomyces cerevisiae", "Saccharomyces", "fungus", 2.0, "gDNA_percent"),
    ],
    "D6300": [
        ("Pseudomonas aeruginosa", "Pseudomonas", "bacterium", 12.0, "gDNA_percent"),
        ("Escherichia coli", "Escherichia", "bacterium", 12.0, "gDNA_percent"),
        ("Salmonella enterica", "Salmonella", "bacterium", 12.0, "gDNA_percent"),
        ("Lactobacillus fermentum", "Lactobacillus", "bacterium", 12.0, "gDNA_percent"),
        ("Enterococcus faecalis", "Enterococcus", "bacterium", 12.0, "gDNA_percent"),
        ("Staphylococcus aureus", "Staphylococcus", "bacterium", 12.0, "gDNA_percent"),
        ("Listeria monocytogenes", "Listeria", "bacterium", 12.0, "gDNA_percent"),
        ("Bacillus subtilis", "Bacillus", "bacterium", 12.0, "gDNA_percent"),
        ("Saccharomyces cerevisiae", "Saccharomyces", "fungus", 2.0, "gDNA_percent"),
        ("Cryptococcus neoformans", "Cryptococcus", "fungus", 2.0, "gDNA_percent"),
    ],
}
GENUS_ALIASES = {
    "Escherichia": {"Escherichia", "Escherichia-Shigella"},
    "Lactobacillus": {"Lactobacillus", "Limosilactobacillus"},
}

# Product-to-trip assignments are author-confirmed, but raw-library identities
# still have to agree with the frozen samplesheet/metadata.  In particular,
# e0323_Ctrl_1_Trip1 is labelled negative in controls-metadata.tsv, so it is
# not promoted to a positive control from its ambiguous name alone.  The seven
# mappings below are evaluation hypotheses from positive labels plus trip-level
# design, not confirmed library-to-product identities.
POSITIVE_16S = {
    "e0553_Ctrl_2": ("Trip1", "D6322", "PCR/library/sequencing"),
    "e0554_Ctrl_3": ("Trip1", "D6322", "PCR/library/sequencing"),
    "e0872_TMC1": ("Trip2", "D6322", "PCR/library/sequencing"),
    "e0874_Positive": ("Trip2", "D6322", "PCR/library/sequencing"),
    "e8294_FPosCtrl1": ("Trip3", "D6300", "extraction/lysis/PCR/library/sequencing"),
    "e8661_FPosCtrl2": ("Trip3", "D6300", "extraction/lysis/PCR/library/sequencing"),
    "e8665_FPosCtrl3": ("Trip3", "D6300", "extraction/lysis/PCR/library/sequencing"),
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def provenance_path(path: Path, root: Path) -> str:
    """Return a stable manifest label without resolving staged symlinks.

    Workflow inputs may be passed as relative paths whose final component is a
    symlink to another process output.  Resolving such a path would move it
    outside ``root`` and make ``relative_to`` fail, even though the staged path
    itself is inside the task sandbox.  ``Path.absolute`` normalizes the
    lexical path while preserving that staging boundary.
    """
    absolute_path = path.absolute()
    absolute_root = root.absolute()
    try:
        return absolute_path.relative_to(absolute_root).as_posix()
    except ValueError:
        return path.as_posix()


def split_ids(value: object) -> list[str]:
    if value is None:
        return []
    return [part.strip() for part in str(value).split(",") if part.strip()]


def parse_trip5_batches(path: Path) -> tuple[dict[str, dict[str, object]], dict[str, str]]:
    ws = load_workbook(path, read_only=True, data_only=True).active
    batches: dict[str, dict[str, object]] = {}
    sample_to_blank: dict[str, str] = {}
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        label = str(row[0]).strip() if row[0] else ""
        if not label.startswith("EB") or not label[2:].isdigit():
            continue
        number = int(label[2:])
        if not 1 <= number <= 17:
            continue
        samples = split_ids(row[4])
        index_code = str(row[2]).strip().replace("  ", " ") if row[2] else ""
        date = datetime.strptime(str(row[1]), "%m/%d/%Y").date().isoformat()
        if int(row[3]) != len(samples):
            raise ValueError(f"{label}: declared {row[3]} samples but parsed {len(samples)}")
        for sample in samples:
            if sample in sample_to_blank:
                raise ValueError(f"{sample} appears in two Trip-5 extraction batches")
            sample_to_blank[sample] = label
        batches[label] = {
            "blank": label,
            "date": date,
            "index_code": index_code,
            "samples": samples,
            "source_row": row_number,
        }
    if set(batches) != {f"EB{i}" for i in range(1, 18)}:
        raise ValueError("Trip-5 workbook must contain exactly EB1-EB17")
    return batches, sample_to_blank


def load_taxonomy(path: Path) -> dict[str, str]:
    result: dict[str, str] = {}
    with path.open(encoding="utf-8") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        id_key = reader.fieldnames[0]
        tax_key = "Taxon" if "Taxon" in reader.fieldnames else reader.fieldnames[1]
        for row in reader:
            result[row[id_key]] = row[tax_key]
    return result


def body(sample_id: str) -> str:
    if sample_id.startswith("e") and "_" in sample_id:
        prefix, rest = sample_id.split("_", 1)
        if prefix[1:].isdigit():
            return rest
    return sample_id


def write_rows(path: Path, fieldnames: list[str], rows: list[dict[str, object]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def positive_control_audit(
    biom_path: Path, taxonomy_path: Path, output: Path
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    table = load_table(str(biom_path))
    sample_ids = set(table.ids(axis="sample"))
    taxonomy = load_taxonomy(taxonomy_path)
    profile_rows: list[dict[str, object]] = []
    recovery_rows: list[dict[str, object]] = []
    dominant_rows: list[dict[str, object]] = []
    for sample_id, (trip, product, evaluable_stage) in POSITIVE_16S.items():
        if sample_id not in sample_ids:
            total = 0
            counts: dict[str, float] = {}
            asvs = 0
            status = "absent_after_denoising"
        else:
            vector = table.data(sample_id, axis="sample", dense=True)
            observation_ids = table.ids(axis="observation")
            total = int(vector.sum())
            asvs = int((vector > 0).sum())
            counts = defaultdict(float)
            lineages = defaultdict(float)
            for feature, count in zip(observation_ids, vector, strict=True):
                if count <= 0:
                    continue
                lineage = taxonomy.get(feature, "")
                lineages[lineage] += float(count)
                parts = lineage.split(";")
                genus = parts[5].strip() if len(parts) > 5 else ""
                if genus and genus not in {"NA", "__"}:
                    counts[genus] += float(count)
            for rank, (lineage, count) in enumerate(
                sorted(lineages.items(), key=lambda item: item[1], reverse=True)[:20],
                start=1,
            ):
                dominant_rows.append(
                    {
                        "profile_id": sample_id,
                        "rank": rank,
                        "taxonomy": lineage,
                        "reads": int(count),
                        "read_fraction": f"{count / total:.8f}" if total else "",
                    }
                )
            status = "profiled"
        expected_bacterial_genera = [
            genus
            for _species, genus, group, _expected, _basis in EXPECTED[product]
            if group == "bacterium"
        ]
        expected_bacterial_reads = sum(
            sum(
                counts.get(alias, 0.0)
                for alias in GENUS_ALIASES.get(genus, {genus})
            )
            for genus in expected_bacterial_genera
        )
        expected_bacterial_recovered = sum(
            any(
                counts.get(alias, 0.0) > 0
                for alias in GENUS_ALIASES.get(genus, {genus})
            )
            for genus in expected_bacterial_genera
        )
        expected_bacterial_fraction = (
            expected_bacterial_reads / total if total else None
        )
        profile_rows.append(
            {
                "profile_id": sample_id,
                "trip": trip,
                "product": product,
                "product_assignment_status": (
                    "provisional_from_positive_label_and_trip_design"
                ),
                "product_form": "purified_HMW_DNA" if product == "D6322" else "whole_cells",
                "evaluable_stage": evaluable_stage,
                "role_in_contaminant_model": "evaluation_only",
                "reads": total,
                "observed_asvs": asvs,
                "expected_bacterial_genera": len(expected_bacterial_genera),
                "expected_bacterial_genera_recovered": expected_bacterial_recovered,
                "expected_bacterial_genus_reads": int(expected_bacterial_reads),
                "expected_bacterial_genus_read_fraction": (
                    f"{expected_bacterial_fraction:.8f}"
                    if expected_bacterial_fraction is not None
                    else ""
                ),
                "profile_interpretation": (
                    "mock_dominated_at_or_above_0.95_expected_bacterial_genus_reads"
                    if expected_bacterial_fraction is not None
                    and expected_bacterial_fraction >= 0.95
                    else "not_mock_dominated_under_declared_0.95_descriptive_threshold"
                ),
                "status": status,
            }
        )
        for species, genus, organism_group, expected, basis in EXPECTED[product]:
            observed = sum(counts.get(alias, 0.0) for alias in GENUS_ALIASES.get(genus, {genus}))
            if organism_group == "fungus":
                call = "not_targeted_by_bacterial_16S"
            elif total == 0:
                call = "not_evaluable_no_post_denoising_profile"
            elif observed > 0:
                call = "recovered_at_genus_level"
            else:
                call = "not_recovered_at_genus_level"
            recovery_rows.append(
                {
                    "profile_id": sample_id,
                    "trip": trip,
                    "product": product,
                    "product_assignment_status": (
                        "provisional_from_positive_label_and_trip_design"
                    ),
                    "expected_species": species,
                    "comparison_rank": "genus",
                    "comparison_name": genus,
                    "organism_group": organism_group,
                    "manufacturer_expected_value": expected,
                    "manufacturer_expected_basis": basis,
                    "observed_reads_at_genus": int(observed),
                    "observed_read_fraction": f"{observed / total:.8f}" if total else "",
                    "recovery_call": call,
                    "quantitative_interpretation": (
                        "descriptive_only; gDNA proportions are not expected 16S read proportions"
                    ),
                }
            )
    write_rows(output / "positive_control_profiles.tsv", list(profile_rows[0]), profile_rows)
    write_rows(output / "positive_control_expected_taxon_recovery.tsv", list(recovery_rows[0]), recovery_rows)
    write_rows(
        output / "positive_control_dominant_taxa.tsv",
        list(dominant_rows[0]) if dominant_rows else [
            "profile_id", "rank", "taxonomy", "reads", "read_fraction"
        ],
        dominant_rows,
    )
    return profile_rows, recovery_rows


def trip5_screen(
    project_root: Path,
    table_path: Path,
    taxonomy_path: Path,
    workbook_path: Path,
    profile_metadata_path: Path,
    output: Path,
    primary_threshold: float,
) -> dict[str, object]:
    batches, sample_to_blank = parse_trip5_batches(workbook_path)
    with table_path.open(encoding="utf-8") as handle:
        provenance_line = handle.readline().rstrip("\n")
        header = handle.readline().rstrip("\n").split("\t")
        columns = header[1:]
        index = {name: i for i, name in enumerate(columns)}
        controls = [f"EB{i}" for i in range(1, 18)]
        missing_controls = sorted(set(controls) - set(index))
        mapped_present = sorted(set(sample_to_blank) & set(index))
        mapped_missing = sorted(set(sample_to_blank) - set(index))
        if missing_controls:
            raise ValueError(f"mapped extraction blanks absent from canonical table: {missing_controls}")
        if not mapped_present:
            raise ValueError("no mapped biological profiles found")
        control_idx = np.array([index[name] for name in controls], dtype=int)
        biological_idx = np.array([index[name] for name in mapped_present], dtype=int)
        feature_ids: list[str] = []
        control_presence: list[int] = []
        biological_presence: list[int] = []
        control_reads: list[float] = []
        biological_reads: list[float] = []
        per_column_total = np.zeros(len(columns), dtype=np.float64)
        per_column_features = np.zeros(len(columns), dtype=np.int64)
        for line in handle:
            fields = line.rstrip("\n").split("\t")
            values = np.fromiter((float(v) for v in fields[1:]), dtype=np.float64, count=len(columns))
            present = values > 0
            feature_ids.append(fields[0])
            control_presence.append(int(present[control_idx].sum()))
            biological_presence.append(int(present[biological_idx].sum()))
            control_reads.append(float(values[control_idx].sum()))
            biological_reads.append(float(values[biological_idx].sum()))
            per_column_total += values
            per_column_features += present

    cp = np.asarray(control_presence)
    bp = np.asarray(biological_presence)
    scores = np.ones(len(feature_ids), dtype=np.float64)
    for i in np.flatnonzero(cp):
        scores[i] = fisher_exact(
            [[int(cp[i]), len(controls) - int(cp[i])],
             [int(bp[i]), len(mapped_present) - int(bp[i])]],
            alternative="greater",
        ).pvalue
    control_prev = cp / len(controls)
    biological_prev = bp / len(mapped_present)
    scenarios: list[dict[str, object]] = []
    primary_called: np.ndarray | None = None
    for threshold in (0.01, 0.05, 0.10):
        for minimum_blanks in (1, 2, 3):
            called = (
                (scores < threshold)
                & (cp >= minimum_blanks)
                & (control_prev > biological_prev)
            )
            scenarios.append(
                {
                    "prevalence_score_threshold": threshold,
                    "minimum_blank_prevalence_count": minimum_blanks,
                    "called_features": int(called.sum()),
                    "control_reads_in_called_features": int(np.asarray(control_reads)[called].sum()),
                    "biological_reads_in_called_features": int(
                        np.asarray(biological_reads)[called].sum()
                    ),
                }
            )
            if threshold == primary_threshold and minimum_blanks == 2:
                primary_called = called
    if primary_called is None:
        raise ValueError("primary threshold must be one of 0.01, 0.05, or 0.10")
    taxonomy = load_taxonomy(taxonomy_path)
    score_rows = []
    for i, feature in enumerate(feature_ids):
        if not primary_called[i]:
            continue
        score_rows.append(
            {
                "feature_id": feature,
                "taxonomy": taxonomy.get(feature, ""),
                "extraction_blanks_present": int(cp[i]),
                "extraction_blanks_total": len(controls),
                "mapped_biological_profiles_present": int(bp[i]),
                "mapped_biological_profiles_total": len(mapped_present),
                "blank_prevalence": f"{control_prev[i]:.8f}",
                "biological_prevalence": f"{biological_prev[i]:.8f}",
                "prevalence_score": f"{scores[i]:.10g}",
                "blank_reads": int(control_reads[i]),
                "mapped_biological_reads": int(biological_reads[i]),
                "call": "candidate_contaminant",
            }
        )
    score_rows.sort(key=lambda row: (float(row["prevalence_score"]), -int(row["blank_reads"])))
    write_rows(
        output / "trip5_primary_contaminant_calls.tsv",
        list(score_rows[0]) if score_rows else [
            "feature_id", "taxonomy", "extraction_blanks_present", "extraction_blanks_total",
            "mapped_biological_profiles_present", "mapped_biological_profiles_total",
            "blank_prevalence", "biological_prevalence", "prevalence_score", "blank_reads",
            "mapped_biological_reads", "call",
        ],
        score_rows,
    )
    write_rows(output / "trip5_filter_sensitivity.tsv", list(scenarios[0]), scenarios)

    called_ids = {feature_ids[i] for i in np.flatnonzero(primary_called)}
    removal_reads = np.zeros(len(columns), dtype=np.float64)
    removal_features = np.zeros(len(columns), dtype=np.int64)
    with table_path.open(encoding="utf-8") as handle:
        handle.readline()
        handle.readline()
        for line in handle:
            fields = line.rstrip("\n").split("\t")
            if fields[0] not in called_ids:
                continue
            values = np.fromiter((float(v) for v in fields[1:]), dtype=np.float64, count=len(columns))
            removal_reads += values
            removal_features += values > 0
    mapped_set = set(mapped_present)
    all_negative = {f"EB{i}" for i in range(1, 19)} | {
        "Negative1", "Negative2", "Negative4", "Negative5", "Negative6", "Negative7"
    }
    removal_rows = []
    for i, name in enumerate(columns):
        if name not in mapped_set and name not in all_negative:
            continue
        total = per_column_total[i]
        role = (
            "training_extraction_blank"
            if name in set(controls)
            else "characterization_only_extraction_blank"
            if name in all_negative
            else "compatible_biological_profile"
        )
        removal_rows.append(
            {
                "profile_id": name,
                "role": role,
                "extraction_blank": sample_to_blank.get(name, ""),
                "total_reads": int(total),
                "total_asvs": int(per_column_features[i]),
                "candidate_contaminant_reads": int(removal_reads[i]),
                "candidate_contaminant_asvs": int(removal_features[i]),
                "candidate_contaminant_read_fraction": f"{removal_reads[i] / total:.8f}" if total else "",
                "application_scope": (
                    "removed_in_trip5_sensitivity_table"
                    if name in mapped_set
                    else "characterized_not_filtered"
                ),
            }
        )
    write_rows(output / "trip5_removal_fraction_by_profile.tsv", list(removal_rows[0]), removal_rows)

    with profile_metadata_path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        profile_key = reader.fieldnames[0]
        profile_metadata = {row[profile_key]: row for row in reader}
    missing_metadata = sorted(mapped_set - set(profile_metadata))
    if missing_metadata:
        raise ValueError(
            "mapped biological profiles absent from profile metadata: "
            + ", ".join(missing_metadata)
        )

    biological_rows = [
        row for row in removal_rows if row["role"] == "compatible_biological_profile"
    ]
    group_outputs = (
        ("campaign", "Trip", "trip5_removal_fraction_by_campaign.tsv"),
        ("compartment", "Type", "trip5_removal_fraction_by_compartment.tsv"),
    )
    for group_type, metadata_field, filename in group_outputs:
        grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
        for row in biological_rows:
            group_value = profile_metadata[str(row["profile_id"])][metadata_field]
            if not group_value:
                raise ValueError(
                    f"{row['profile_id']}: missing {metadata_field} in profile metadata"
                )
            grouped[group_value].append(row)
        group_rows = []
        for group_value, member_rows in sorted(grouped.items()):
            total_reads = sum(int(row["total_reads"]) for row in member_rows)
            removed_reads = sum(
                int(row["candidate_contaminant_reads"]) for row in member_rows
            )
            fractions = [
                float(row["candidate_contaminant_read_fraction"])
                for row in member_rows
            ]
            group_rows.append(
                {
                    "group_type": group_type,
                    "group_value": group_value,
                    "biological_profile_count": len(member_rows),
                    "total_reads": total_reads,
                    "candidate_contaminant_reads": removed_reads,
                    "pooled_candidate_contaminant_read_fraction": (
                        f"{removed_reads / total_reads:.8f}" if total_reads else ""
                    ),
                    "median_candidate_contaminant_read_fraction": (
                        f"{np.median(fractions):.8f}" if fractions else ""
                    ),
                    "maximum_candidate_contaminant_read_fraction": (
                        f"{max(fractions):.8f}" if fractions else ""
                    ),
                }
            )
        write_rows(output / filename, list(group_rows[0]), group_rows)

    batch_rows = []
    by_name = {row["profile_id"]: row for row in removal_rows}
    for blank, record in sorted(batches.items(), key=lambda item: int(item[0][2:])):
        present_samples = [s for s in record["samples"] if s in by_name]
        fractions = [
            float(by_name[s]["candidate_contaminant_read_fraction"])
            for s in present_samples
            if by_name[s]["candidate_contaminant_read_fraction"] != ""
        ]
        batch_rows.append(
            {
                "extraction_blank": blank,
                "extraction_date": record["date"],
                "index_code": record["index_code"],
                "workbook_sample_count": len(record["samples"]),
                "canonical_sample_count": len(present_samples),
                "canonical_missing_sample_count": len(record["samples"]) - len(present_samples),
                "blank_reads": by_name[blank]["total_reads"],
                "blank_asvs": by_name[blank]["total_asvs"],
                "median_biological_removal_fraction": f"{np.median(fractions):.8f}" if fractions else "",
                "maximum_biological_removal_fraction": f"{max(fractions):.8f}" if fractions else "",
            }
        )
    write_rows(output / "trip5_extraction_batch_summary.tsv", list(batch_rows[0]), batch_rows)

    # Replayable sensitivity-analysis input: the exact mapped Trip-5
    # biological universe with primary candidate features removed.  This does
    # not replace the canonical table and deliberately excludes controls and
    # unmapped Trip-5 profiles.
    mapped_positions = [index[name] for name in mapped_present]
    filtered_path = output / "trip5_mapped_feature_table_control_filtered.tsv.gz"
    retained_features = 0
    with (
        table_path.open(encoding="utf-8") as source_handle,
        filtered_path.open("wb") as raw_handle,
        gzip.GzipFile(
            filename="", fileobj=raw_handle, mode="wb", mtime=0
        ) as gzip_handle,
        io.TextIOWrapper(
            gzip_handle, encoding="utf-8", newline=""
        ) as filtered_handle,
    ):
        source_handle.readline()
        source_handle.readline()
        filtered_handle.write(
            "# Control-sensitivity table; source="
            + str(table_path)
            + f"; threshold={primary_threshold}; minimum_blanks=2; "
            + "training_controls=EB1-EB17\n"
        )
        filtered_handle.write("#OTU ID\t" + "\t".join(mapped_present) + "\n")
        for line in source_handle:
            fields = line.rstrip("\n").split("\t")
            if fields[0] in called_ids:
                continue
            selected = [fields[position + 1] for position in mapped_positions]
            if not any(float(value) > 0 for value in selected):
                continue
            filtered_handle.write(fields[0] + "\t" + "\t".join(selected) + "\n")
            retained_features += 1

    training_rows = []
    for blank in controls:
        training_rows.append(
            {
                "profile_id": blank,
                "control_role": "extraction_blank",
                "assay": "16S_V3-V4",
                "use": "contaminant_training",
                "extraction_date": batches[blank]["date"],
                "source": "Sequenced_Samples_by_EB_FifthTrip.xlsx",
            }
        )
    for name in sorted(all_negative - set(controls)):
        if name in index:
            training_rows.append(
                {
                    "profile_id": name,
                    "control_role": "extraction_blank",
                    "assay": "16S_V3-V4",
                    "use": "characterization_only_unmapped_extraction_batch",
                    "extraction_date": "",
                    "source": "ibex_trip5_16s_samplesheet.tsv",
                }
            )
    for name in POSITIVE_16S:
        training_rows.append(
            {
                "profile_id": name,
                "control_role": "positive_control",
                "assay": "16S_V3-V4",
                "use": "evaluation_only_never_contaminant_training",
                "extraction_date": "",
                "source": "ibex_20250714_qiime2/table.qza",
            }
        )
    contaminant_training_profiles = {
        row["profile_id"]
        for row in training_rows
        if row["use"] == "contaminant_training"
    }
    positive_profiles = {
        row["profile_id"]
        for row in training_rows
        if row["control_role"] == "positive_control"
    }
    positive_training_overlap = sorted(
        contaminant_training_profiles & positive_profiles
    )
    if positive_training_overlap:
        raise ValueError(
            "positive controls entered contaminant training: "
            + ", ".join(positive_training_overlap)
        )
    write_rows(output / "control_analysis_roles.tsv", list(training_rows[0]), training_rows)

    try:
        filtered_display = str(filtered_path.relative_to(project_root))
    except ValueError:
        filtered_display = str(filtered_path)

    return {
        "canonical_table_provenance_line": provenance_line,
        "canonical_profiles": len(columns),
        "canonical_features": len(feature_ids),
        "training_extraction_blanks": controls,
        "characterization_only_negative_profiles": sorted(all_negative - set(controls)),
        "mapped_biological_profiles_in_workbook": len(sample_to_blank),
        "mapped_biological_profiles_in_canonical_table": len(mapped_present),
        "mapped_biological_profiles_absent_from_canonical_table": mapped_missing,
        "primary_method": "one-sided Fisher prevalence enrichment",
        "primary_prevalence_score_threshold": primary_threshold,
        "primary_minimum_blank_prevalence_count": 2,
        "primary_candidate_contaminant_features": int(primary_called.sum()),
        "filtered_sensitivity_table": filtered_display,
        "filtered_sensitivity_table_profiles": len(mapped_present),
        "filtered_sensitivity_table_retained_features": retained_features,
        "filtered_sensitivity_table_sha256": sha256(filtered_path),
        "frequency_method": "not_run; DNA concentration data unavailable for these controls",
        "batch_method": (
            "pooled prevalence inference across EB1-EB17; a single candidate "
            "set is applied uniformly to the union of their mapped "
            "extraction-batch samples, and removal is summarized per batch"
        ),
        "removal_summary_dimensions": [
            "profile",
            "campaign",
            "compartment",
            "extraction_batch",
        ],
        "positive_library_product_assignment": (
            "provisional_from_positive_label_and_trip_design"
        ),
        "positive_controls_in_training": len(positive_training_overlap),
        "limitations": [
            "one extraction blank per extraction day prevents within-batch prevalence estimation",
            "EB18 and Negative1/2/4-7 lack extraction-day mappings and are characterization-only",
            "Trip-4 blank FASTQs are not present in the canonical denoising table",
            "Trip-5 D6300 positive control is shotgun-only and cannot validate the 16S assay",
            "candidate calls are a sensitivity analysis and do not replace the unfiltered canonical table",
        ],
    }


def main() -> int:
    root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--canonical-table",
        type=Path,
        default=root / "data/processed/taxonomy/taxon-tables/feature-table-trips1-5.tsv",
    )
    parser.add_argument(
        "--canonical-taxonomy",
        type=Path,
        default=root / "data/processed/taxonomy/taxon-tables/taxonomy-trips1-5.tsv",
    )
    parser.add_argument(
        "--trip5-workbook",
        type=Path,
        default=root / "data/metadata/samples/Sequenced_Samples_by_EB_FifthTrip.xlsx",
    )
    parser.add_argument(
        "--profile-metadata",
        type=Path,
        default=root / "analysis/v3/figure_inputs/cache/alpha.tsv",
    )
    parser.add_argument(
        "--july-biom",
        type=Path,
        default=root
        / "data/metadata/samples/controls/source_snapshots/ibex_20250714_qiime2/extracted/feature-table.biom",
    )
    parser.add_argument(
        "--july-taxonomy",
        type=Path,
        default=root
        / "data/metadata/samples/controls/source_snapshots/ibex_20250714_qiime2/extracted/taxonomy.tsv",
    )
    parser.add_argument(
        "--output-dir", type=Path, default=root / "analysis/v3/control_audit"
    )
    parser.add_argument("--primary-threshold", type=float, default=0.10)
    args = parser.parse_args()
    inputs = [
        args.canonical_table,
        args.canonical_taxonomy,
        args.trip5_workbook,
        args.profile_metadata,
        args.july_biom,
        args.july_taxonomy,
    ]
    missing = [str(path) for path in inputs if not path.is_file()]
    if missing:
        raise FileNotFoundError("missing required inputs: " + ", ".join(missing))
    args.output_dir.mkdir(parents=True, exist_ok=True)
    profiles, recovery = positive_control_audit(
        args.july_biom, args.july_taxonomy, args.output_dir
    )
    summary = trip5_screen(
        root,
        args.canonical_table,
        args.canonical_taxonomy,
        args.trip5_workbook,
        args.profile_metadata,
        args.output_dir,
        args.primary_threshold,
    )
    summary.update(
        {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "method_reference": "Fierer et al. 2025, doi:10.1038/s41564-025-02035-2",
            "input_sha256": {
                provenance_path(path, root): sha256(path) for path in inputs
            },
            "positive_control_profiles": len(profiles),
            "positive_expected_taxon_evaluations": len(recovery),
        }
    )
    with (args.output_dir / "summary.json").open("w", encoding="utf-8") as handle:
        json.dump(summary, handle, indent=2, sort_keys=True)
        handle.write("\n")
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
